import argparse
import gc
import json
import os
import re
import unicodedata
from collections import Counter, defaultdict
from itertools import combinations

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import networkx as nx
import numpy as np
import pandas as pd
import seaborn as sns
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

plt.style.use("seaborn-v0_8-whitegrid")
sns.set_palette("viridis")


def clean_text(value):
    if pd.isna(value):
        return ""
    value = str(value)
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def safe_filename(value):
    value = clean_text(value)
    value = value.replace("/", "_").replace("\\", "_")
    value = re.sub(r"[^A-Za-z0-9_. -]+", "", value)
    value = value.replace(" ", "_")
    value = re.sub(r"_+", "_", value)
    return value.strip("_")


def remove_existing_files(paths):
    for path in paths:
        if os.path.exists(path):
            os.remove(path)


def read_preprocessed(path):
    df = pd.read_csv(path, dtype=str, low_memory=False).fillna("")
    if "Original CSV Row" not in df.columns:
        df.insert(0, "Original CSV Row", np.arange(2, len(df) + 2))
    if "Cited by" not in df.columns:
        df["Cited by"] = 0
    if "Year" not in df.columns:
        df["Year"] = ""
    df["Cited by"] = pd.to_numeric(df["Cited by"], errors="coerce").fillna(0).astype(int)
    df["Year"] = pd.to_numeric(df["Year"], errors="coerce")
    for col in ["Author full names", "Affiliations", "Source title", "Editors", "Publisher", "Language of Original Document", "Open Access", "Title", "Abstract", "Author Keywords", "Index Keywords"]:
        if col not in df.columns:
            df[col] = ""
        df[col] = df[col].fillna("").astype(str)
    return df


def list_columns(df, prefix):
    cols = [col for col in df.columns if re.fullmatch(fr"{prefix}_\d+", str(col))]
    return sorted(cols, key=lambda x: int(x.split("_")[-1]))


def entities_per_row(df, prefix):
    cols = list_columns(df, prefix)
    if not cols:
        raise ValueError(
            f'No {prefix}_# columns were found in the preprocessed input. '
            'Run preprocess_scopus_bibliometrics.py before running the analysis.'
        )
    output = []
    for _, row in df.iterrows():
        values = []
        for col in cols:
            value = clean_text(row.get(col, ""))
            if value:
                values.append(value)
        output.append(list(dict.fromkeys(values)))
    return output


def flatten_unique(entity_lists):
    return pd.Series([entity for entities in entity_lists for entity in list(dict.fromkeys([x for x in entities if x]))])


def productivity_and_citations(df, entity_lists, label):
    publications = Counter()
    citations = Counter()
    normalized_citations = Counter()
    normalized_publications = Counter()
    has_normalized = "Citations per Year" in df.columns
    for row_index, entities in enumerate(entity_lists):
        citation = df.iloc[row_index]["Cited by"]
        normalized_value = df.iloc[row_index]["Citations per Year"] if has_normalized else np.nan
        for entity in set([x for x in entities if x]):
            publications[entity] += 1
            citations[entity] += int(citation)
            if has_normalized and pd.notna(normalized_value):
                normalized_citations[entity] += float(normalized_value)
                normalized_publications[entity] += 1
    rows = []
    for entity in publications:
        normalized_total = normalized_citations[entity]
        normalized_count = normalized_publications[entity]
        rows.append({
            label: entity,
            "Publications": publications[entity],
            "Citations": citations[entity],
            "Publications Included in Annualized Citation Analysis": normalized_count,
            "Total Annualized Citation Rate": normalized_total,
            "Mean Annualized Citation Rate per Paper": normalized_total / normalized_count if normalized_count else 0
        })
    if not rows:
        return pd.DataFrame(columns=[label, "Publications", "Citations", "Publications Included in Annualized Citation Analysis", "Total Annualized Citation Rate", "Mean Annualized Citation Rate per Paper"])
    return pd.DataFrame(rows).sort_values(["Publications", "Citations", label], ascending=[False, False, True])


def single_column_productivity_and_citations(df, column, label):
    publications = Counter()
    citations = Counter()
    normalized_citations = Counter()
    normalized_publications = Counter()
    has_normalized = "Citations per Year" in df.columns
    for _, row in df.iterrows():
        entity = clean_text(row.get(column, ""))
        if entity:
            publications[entity] += 1
            citations[entity] += int(row["Cited by"])
            normalized_value = row.get("Citations per Year", np.nan) if has_normalized else np.nan
            if has_normalized and pd.notna(normalized_value):
                normalized_citations[entity] += float(normalized_value)
                normalized_publications[entity] += 1
    rows = []
    for entity in publications:
        normalized_total = normalized_citations[entity]
        normalized_count = normalized_publications[entity]
        rows.append({
            label: entity,
            "Publications": publications[entity],
            "Citations": citations[entity],
            "Publications Included in Annualized Citation Analysis": normalized_count,
            "Total Annualized Citation Rate": normalized_total,
            "Mean Annualized Citation Rate per Paper": normalized_total / normalized_count if normalized_count else 0
        })
    if not rows:
        return pd.DataFrame(columns=[label, "Publications", "Citations", "Publications Included in Annualized Citation Analysis", "Total Annualized Citation Rate", "Mean Annualized Citation Rate per Paper"])
    return pd.DataFrame(rows).sort_values(["Publications", "Citations", label], ascending=[False, False, True])


def format_plot_value(value, value_format="int"):
    if pd.isna(value):
        return ""
    if value_format == "float":
        return f"{float(value):.2f}".rstrip("0").rstrip(".")
    return str(int(round(float(value))))


def save_horizontal_bar_from_series(series, title, xlabel, ylabel, output_path, n=30):
    series = series.replace("", np.nan).dropna() if isinstance(series, pd.Series) else series
    if isinstance(series, pd.Series):
        plot_data = series.value_counts().nlargest(n)
    else:
        plot_data = pd.Series(dtype=float)
    if plot_data.empty:
        return
    plt.figure(figsize=(12, 10))
    ax = sns.barplot(x=plot_data.values, y=plot_data.index, hue=plot_data.index, palette=sns.color_palette("viridis", n_colors=len(plot_data)), legend=False)
    plt.ylabel(ylabel, fontsize=12)
    plt.xlabel(xlabel, fontsize=12)
    for i, v in enumerate(plot_data.values):
        ax.text(v / 2, i, str(int(v)), color="white", fontweight="bold", ha="center", va="center", fontsize=10)
    plt.title(title, fontsize=16, weight="bold")
    plt.tight_layout()
    plt.savefig(output_path, dpi=600, format="jpeg", bbox_inches="tight")
    plt.close()


def save_horizontal_bar_from_table(table, value_col, label_col, title, xlabel, ylabel, output_path, n=30, label_suffix_col=None, sort_value_col=None, value_format="int"):
    if table.empty:
        return
    sort_col = sort_value_col if sort_value_col else value_col
    plot_data = table.sort_values([sort_col, value_col, label_col], ascending=[False, False, True]).head(n).copy()
    if label_suffix_col and label_suffix_col in plot_data.columns:
        plot_data[label_col] = plot_data.apply(lambda row: f'{row[label_col]} ({int(row[label_suffix_col])} papers)', axis=1)
    plt.figure(figsize=(12, 10))
    ax = sns.barplot(x=plot_data[value_col].values, y=plot_data[label_col].values, hue=plot_data[label_col].values, palette=sns.color_palette("viridis", n_colors=len(plot_data)), legend=False)
    plt.ylabel(ylabel, fontsize=12)
    plt.xlabel(xlabel, fontsize=12)
    for i, v in enumerate(plot_data[value_col].values):
        ax.text(v / 2 if v != 0 else 0, i, format_plot_value(v, value_format), color="white" if v != 0 else "black", fontweight="bold", ha="center", va="center", fontsize=10)
    plt.title(title, fontsize=16, weight="bold")
    plt.tight_layout()
    plt.savefig(output_path, dpi=600, format="jpeg", bbox_inches="tight")
    plt.close()


def save_vertical_count_bar(series, title, xlabel, ylabel, output_path, n=10):
    series = series.replace("", np.nan).dropna()
    if series.empty:
        return
    plot_data = series.value_counts().nlargest(n)
    if plot_data.empty:
        return
    plt.figure(figsize=(12, 10))
    ax = sns.barplot(x=plot_data.index, y=plot_data.values, hue=plot_data.index, palette=sns.color_palette("plasma", n_colors=len(plot_data)), legend=False)
    plt.xlabel(xlabel, fontsize=12)
    plt.ylabel(ylabel, fontsize=12)
    plt.xticks(rotation=45, ha="right")
    max_value = max(plot_data.values)
    for i, v in enumerate(plot_data.values):
        ax.text(i, v + 0.01 * max_value, str(int(v)), ha="center", va="bottom", fontsize=10)
    plt.title(title, fontsize=16, weight="bold")
    plt.tight_layout()
    plt.savefig(output_path, dpi=600, format="jpeg", bbox_inches="tight")
    plt.close()


def save_year_bar(years, values, title, ylabel, output_path, value_format="int"):
    if len(years) == 0:
        return
    plt.figure(figsize=(14, 7))
    ax = sns.barplot(x=[int(year) for year in years], y=values, hue=[int(year) for year in years], palette=sns.color_palette("plasma", n_colors=len(years)), legend=False)
    plt.title(title, fontsize=16, weight="bold")
    plt.xlabel("Year", fontsize=12)
    plt.ylabel(ylabel, fontsize=12)
    plt.xticks(rotation=45)
    max_value = max(values) if len(values) else 0
    for i, v in enumerate(values):
        ax.text(i, v + 0.01 * max_value if max_value else v, format_plot_value(v, value_format), ha="center", va="bottom", fontsize=8)
    plt.tight_layout()
    plt.savefig(output_path, dpi=600, format="jpeg", bbox_inches="tight")
    plt.close()


def normalize_open_access_status(value):
    normalized = clean_text(value).casefold()
    normalized = re.sub(r"[^a-z0-9]+", " ", normalized).strip()
    return "Open Access" if "all open access" in normalized else "Not Open Access"


def save_open_access_pie(df, title, output_path):
    status = df["Open Access"].apply(normalize_open_access_status)
    counts = status.value_counts()
    counts = counts.reindex(["Open Access", "Not Open Access"]).dropna()
    counts.to_frame("Publications").to_csv(output_path.replace(".jpeg", ".csv"), encoding="utf-8-sig")
    if counts.empty:
        return
    plt.figure(figsize=(12, 10))
    plt.pie(counts.values, labels=counts.index, autopct="%1.1f%%", startangle=140, colors=sns.color_palette("plasma", n_colors=len(counts)))
    plt.ylabel("")
    plt.title(title, fontsize=16, weight="bold")
    plt.tight_layout()
    plt.savefig(output_path, dpi=600, format="jpeg", bbox_inches="tight")
    plt.close()


def add_open_access_status(df):
    output = df.copy()
    output["Open Access Status"] = output["Open Access"].apply(normalize_open_access_status)
    return output


def mean_or_zero(values):
    values = pd.to_numeric(values, errors="coerce").dropna()
    return float(values.mean()) if len(values) else 0


def ratio_or_nan(numerator, denominator):
    numerator = float(numerator) if pd.notna(numerator) else np.nan
    denominator = float(denominator) if pd.notna(denominator) else np.nan
    if pd.isna(denominator) or denominator == 0:
        return np.nan
    return numerator / denominator


def open_access_status_summary_table(df):
    working = add_open_access_status(df)
    rows = []
    total_publications = len(working)
    for status in ["Open Access", "Not Open Access"]:
        subset = working[working["Open Access Status"] == status].copy()
        citations = pd.to_numeric(subset["Cited by"], errors="coerce").fillna(0)
        normalized = pd.to_numeric(subset["Citations per Year"], errors="coerce") if "Citations per Year" in subset.columns else pd.Series(dtype=float)
        rows.append({
            "Open Access Status": status,
            "Publications": int(len(subset)),
            "Publication Share (%)": float(len(subset) / total_publications * 100) if total_publications else 0,
            "Total Citation Impact": int(citations.sum()) if len(citations) else 0,
            "Mean Cited by per Paper": float(citations.mean()) if len(citations) else 0,
            "Median Cited by per Paper": float(citations.median()) if len(citations) else 0,
            "Publications Included in Annualized Citation Analysis": int(normalized.notna().sum()) if len(normalized) else 0,
            "Total Annualized Citation Rate": float(normalized.sum(skipna=True)) if len(normalized) else 0,
            "Mean Annualized Citation Rate per Paper": float(normalized.mean(skipna=True)) if normalized.notna().any() else 0,
            "Median Annualized Citation Rate per Paper": float(normalized.median(skipna=True)) if normalized.notna().any() else 0
        })
    return pd.DataFrame(rows)


def open_access_unadjusted_comparison_table(summary):
    if summary.empty:
        return pd.DataFrame()
    indexed = summary.set_index("Open Access Status")
    if "Open Access" not in indexed.index or "Not Open Access" not in indexed.index:
        return pd.DataFrame()
    metrics = [
        "Mean Cited by per Paper",
        "Median Cited by per Paper",
        "Mean Annualized Citation Rate per Paper",
        "Median Annualized Citation Rate per Paper"
    ]
    rows = []
    for metric in metrics:
        open_access_value = float(indexed.loc["Open Access", metric])
        non_open_access_value = float(indexed.loc["Not Open Access", metric])
        rows.append({
            "Metric": metric,
            "Open Access": open_access_value,
            "Not Open Access": non_open_access_value,
            "Open Access Minus Not Open Access": open_access_value - non_open_access_value,
            "Open Access to Not Open Access Ratio": ratio_or_nan(open_access_value, non_open_access_value)
        })
    return pd.DataFrame(rows)

def open_access_country_table(df, country_lists):
    working = add_open_access_status(df)
    rows = defaultdict(lambda: {
        "Publications": 0,
        "Open Access Publications": 0,
        "Not Open Access Publications": 0,
        "Total Citation Impact": 0,
        "Open Access Citation Impact": 0,
        "Not Open Access Citation Impact": 0,
        "Open Access Total Annualized Citation Rate": 0.0,
        "Not Open Access Total Annualized Citation Rate": 0.0,
        "Open Access Publications Included in Annualized Citation Analysis": 0,
        "Not Open Access Publications Included in Annualized Citation Analysis": 0
    })
    for row_index, countries in enumerate(country_lists):
        unique_countries = list(dict.fromkeys([country for country in countries if country]))
        if not unique_countries:
            continue
        status = working.iloc[row_index]["Open Access Status"]
        citation = int(working.iloc[row_index]["Cited by"])
        annualized_value = working.iloc[row_index]["Citations per Year"] if "Citations per Year" in working.columns else np.nan
        for country in unique_countries:
            rows[country]["Publications"] += 1
            rows[country]["Total Citation Impact"] += citation
            if status == "Open Access":
                rows[country]["Open Access Publications"] += 1
                rows[country]["Open Access Citation Impact"] += citation
                if pd.notna(annualized_value):
                    rows[country]["Open Access Total Annualized Citation Rate"] += float(annualized_value)
                    rows[country]["Open Access Publications Included in Annualized Citation Analysis"] += 1
            else:
                rows[country]["Not Open Access Publications"] += 1
                rows[country]["Not Open Access Citation Impact"] += citation
                if pd.notna(annualized_value):
                    rows[country]["Not Open Access Total Annualized Citation Rate"] += float(annualized_value)
                    rows[country]["Not Open Access Publications Included in Annualized Citation Analysis"] += 1
    output_rows = []
    for country, values in rows.items():
        publications = values["Publications"]
        open_access_publications = values["Open Access Publications"]
        non_open_access_publications = values["Not Open Access Publications"]
        open_access_mean = values["Open Access Citation Impact"] / open_access_publications if open_access_publications else np.nan
        non_open_access_mean = values["Not Open Access Citation Impact"] / non_open_access_publications if non_open_access_publications else np.nan
        open_access_annualized_mean = values["Open Access Total Annualized Citation Rate"] / values["Open Access Publications Included in Annualized Citation Analysis"] if values["Open Access Publications Included in Annualized Citation Analysis"] else np.nan
        non_open_access_annualized_mean = values["Not Open Access Total Annualized Citation Rate"] / values["Not Open Access Publications Included in Annualized Citation Analysis"] if values["Not Open Access Publications Included in Annualized Citation Analysis"] else np.nan
        output_rows.append({
            "Country": country,
            "Publications": publications,
            "Open Access Publications": open_access_publications,
            "Not Open Access Publications": non_open_access_publications,
            "Open Access Percentage": open_access_publications / publications * 100 if publications else 0,
            "Total Citation Impact": values["Total Citation Impact"],
            "Open Access Citation Impact": values["Open Access Citation Impact"],
            "Not Open Access Citation Impact": values["Not Open Access Citation Impact"],
            "Mean Citations per Open Access Paper": open_access_mean,
            "Mean Citations per Not Open Access Paper": non_open_access_mean,
            "Unadjusted Open Access Minus Not Open Access Mean Citations": open_access_mean - non_open_access_mean if pd.notna(open_access_mean) and pd.notna(non_open_access_mean) else np.nan,
            "Unadjusted Open Access to Not Open Access Mean Citation Ratio": ratio_or_nan(open_access_mean, non_open_access_mean),
            "Open Access Total Annualized Citation Rate": values["Open Access Total Annualized Citation Rate"],
            "Not Open Access Total Annualized Citation Rate": values["Not Open Access Total Annualized Citation Rate"],
            "Mean Annualized Citation Rate per Open Access Paper": open_access_annualized_mean,
            "Mean Annualized Citation Rate per Not Open Access Paper": non_open_access_annualized_mean,
            "Unadjusted Open Access Minus Not Open Access Mean Annualized Citation Rate": open_access_annualized_mean - non_open_access_annualized_mean if pd.notna(open_access_annualized_mean) and pd.notna(non_open_access_annualized_mean) else np.nan,
            "Unadjusted Open Access to Not Open Access Mean Annualized Citation Rate Ratio": ratio_or_nan(open_access_annualized_mean, non_open_access_annualized_mean),
            "Open Access Publications Included in Annualized Citation Analysis": values["Open Access Publications Included in Annualized Citation Analysis"],
            "Not Open Access Publications Included in Annualized Citation Analysis": values["Not Open Access Publications Included in Annualized Citation Analysis"]
        })
    if not output_rows:
        return pd.DataFrame()
    return pd.DataFrame(output_rows).sort_values(["Publications", "Open Access Publications", "Country"], ascending=[False, False, True])

def open_access_annual_table(df):
    working = add_open_access_status(df).dropna(subset=["Year"]).copy()
    if working.empty:
        return pd.DataFrame()
    working["Year"] = working["Year"].astype(int)
    rows = []
    for year, group in working.groupby("Year"):
        open_subset = group[group["Open Access Status"] == "Open Access"]
        non_open_subset = group[group["Open Access Status"] == "Not Open Access"]
        open_annualized = pd.to_numeric(open_subset["Citations per Year"], errors="coerce") if "Citations per Year" in open_subset.columns else pd.Series(dtype=float)
        non_open_annualized = pd.to_numeric(non_open_subset["Citations per Year"], errors="coerce") if "Citations per Year" in non_open_subset.columns else pd.Series(dtype=float)
        rows.append({
            "Year": int(year),
            "Publications": int(len(group)),
            "Open Access Publications": int(len(open_subset)),
            "Not Open Access Publications": int(len(non_open_subset)),
            "Open Access Percentage": float(len(open_subset) / len(group) * 100) if len(group) else 0,
            "Mean Citations per Open Access Paper": mean_or_zero(open_subset["Cited by"]),
            "Mean Citations per Not Open Access Paper": mean_or_zero(non_open_subset["Cited by"]),
            "Mean Annualized Citation Rate per Open Access Paper": float(open_annualized.mean(skipna=True)) if open_annualized.notna().any() else 0,
            "Mean Annualized Citation Rate per Not Open Access Paper": float(non_open_annualized.mean(skipna=True)) if non_open_annualized.notna().any() else 0
        })
    return pd.DataFrame(rows).sort_values("Year")

def descriptive_statistics_values(values):
    numeric = pd.to_numeric(pd.Series(values), errors="coerce").replace([np.inf, -np.inf], np.nan).dropna()
    if numeric.empty:
        return None
    return {
        "Observations": int(numeric.count()),
        "Mean": float(numeric.mean()),
        "Median": float(numeric.median()),
        "Standard Deviation": float(numeric.std(ddof=1)) if len(numeric) > 1 else 0.0,
        "Minimum": float(numeric.min()),
        "Maximum": float(numeric.max())
    }


def add_descriptive_statistics_row(rows, dataset, analysis, unit_of_observation, metric, values):
    stats = descriptive_statistics_values(values)
    if stats is None:
        return
    row = {
        "Dataset": dataset,
        "Analysis": analysis,
        "Unit of Observation": unit_of_observation,
        "Metric": metric
    }
    row.update(stats)
    rows.append(row)

def add_table_descriptive_statistics(rows, dataset, analysis, unit_of_observation, table, metric_columns):
    if table.empty:
        return
    for metric in metric_columns:
        if metric in table.columns:
            add_descriptive_statistics_row(rows, dataset, analysis, unit_of_observation, metric, table[metric])

def descriptive_statistics_table(df, dataset, author_lists, country_lists, author_table, country_table, journal_table, publisher_table, year_counts, year_citation_impact, year_annualized_impact, year_mean_annualized_impact):
    rows = []
    publication_metrics = [
        "Publications",
        "Citations",
        "Publications Included in Annualized Citation Analysis",
        "Total Annualized Citation Rate",
        "Mean Annualized Citation Rate per Paper"
    ]
    add_table_descriptive_statistics(rows, dataset, "Authors", "Unique Author", author_table, publication_metrics)
    add_table_descriptive_statistics(rows, dataset, "Countries", "Unique Country", country_table, publication_metrics)
    add_table_descriptive_statistics(rows, dataset, "Journals", "Unique Journal", journal_table, publication_metrics)
    add_table_descriptive_statistics(rows, dataset, "Publishers", "Unique Publisher", publisher_table, publication_metrics)
    language_table = single_column_productivity_and_citations(df, "Language of Original Document", "Language") if "Language of Original Document" in df.columns else pd.DataFrame()
    document_type_table = single_column_productivity_and_citations(df, "Document Type", "Document Type") if "Document Type" in df.columns else pd.DataFrame()
    add_table_descriptive_statistics(rows, dataset, "Languages", "Language Category", language_table, publication_metrics)
    add_table_descriptive_statistics(rows, dataset, "Document Types", "Document Type Category", document_type_table, publication_metrics)
    working_open_access = add_open_access_status(df)
    open_access_indicator = working_open_access["Open Access Status"].map({"Open Access": 1, "Not Open Access": 0})
    add_descriptive_statistics_row(rows, dataset, "Open Access Status", "Paper", "Open Access Indicator per Paper (Open Access=1)", open_access_indicator)
    add_descriptive_statistics_row(rows, dataset, "Records", "Paper", "Citations per Paper", df["Cited by"])
    if "Citation Years" in df.columns:
        add_descriptive_statistics_row(rows, dataset, "Records", "Paper", "Citation Exposure Years per Paper", df["Citation Years"])
    if "Citations per Year" in df.columns:
        add_descriptive_statistics_row(rows, dataset, "Records", "Paper", "Annualized Citation Rate per Paper", df["Citations per Year"])
    add_descriptive_statistics_row(rows, dataset, "Records", "Paper", "Authors per Paper", [len(values) for values in author_lists])
    add_descriptive_statistics_row(rows, dataset, "Records", "Paper", "Countries per Paper", [len(values) for values in country_lists])
    if not year_counts.empty:
        add_descriptive_statistics_row(rows, dataset, "Publication Years", "Publication Year", "Publications per Year", year_counts.values)
    if not year_citation_impact.empty:
        add_descriptive_statistics_row(rows, dataset, "Publication Years", "Publication Year", "Total Citations by Publication-Year Cohort", year_citation_impact.values)
    if not year_annualized_impact.empty:
        add_descriptive_statistics_row(rows, dataset, "Publication Years", "Publication Year", "Total Annualized Citation Rate by Publication Year", year_annualized_impact.values)
    if not year_mean_annualized_impact.empty:
        add_descriptive_statistics_row(rows, dataset, "Publication Years", "Publication Year", "Mean Annualized Citation Rate per Paper by Publication Year", year_mean_annualized_impact.values)
    return pd.DataFrame(rows)

def save_open_access_extended_analysis(df, country_lists, condition, base_name, results_dir, tables_dir):
    status_summary = open_access_status_summary_table(df)
    unadjusted_comparison = open_access_unadjusted_comparison_table(status_summary)
    country_open_access = open_access_country_table(df, country_lists)
    annual_open_access = open_access_annual_table(df)
    status_summary.to_csv(os.path.join(tables_dir, f"{safe_filename(base_name)}_open_access_citation_impact_summary.csv"), index=False, encoding="utf-8-sig")
    unadjusted_comparison.to_csv(os.path.join(tables_dir, f"{safe_filename(base_name)}_open_access_unadjusted_citation_comparison.csv"), index=False, encoding="utf-8-sig")
    country_open_access.to_csv(os.path.join(tables_dir, f"{safe_filename(base_name)}_open_access_regional_analysis_by_country.csv"), index=False, encoding="utf-8-sig")
    annual_open_access.to_csv(os.path.join(tables_dir, f"{safe_filename(base_name)}_open_access_annual_trends.csv"), index=False, encoding="utf-8-sig")
    save_horizontal_bar_from_table(status_summary, "Mean Cited by per Paper", "Open Access Status", f"Mean Citations per Paper by Open Access Status - {condition}", "Mean Citations per Paper", "Open Access Status", os.path.join(results_dir, f"{safe_filename(base_name)}_Open_Access_Mean_Citations_per_Paper.jpeg"), n=2, value_format="float")
    save_horizontal_bar_from_table(status_summary, "Mean Annualized Citation Rate per Paper", "Open Access Status", f"Mean Annualized Citation Rate per Paper by Open Access Status - {condition}", "Mean Annualized Citation Rate per Paper", "Open Access Status", os.path.join(results_dir, f"{safe_filename(base_name)}_Open_Access_Mean_Annualized_Citation_Rate_per_Paper.jpeg"), n=2, value_format="float")
    if not country_open_access.empty:
        save_horizontal_bar_from_table(country_open_access, "Open Access Publications", "Country", f"Top 30 Countries by Open Access Publication Count - {condition}", "Open Access Papers", "Country", os.path.join(results_dir, f"{safe_filename(base_name)}_Top_30_Countries_by_Open_Access_Publication_Count.jpeg"), n=30)
        rate_plot = country_open_access[country_open_access["Publications"] >= 5].copy()
        rate_plot = rate_plot.dropna(subset=["Open Access Percentage"])
        save_horizontal_bar_from_table(rate_plot, "Open Access Percentage", "Country", f"Top 30 Countries by Open Access Percentage - {condition}", "Open Access Papers (%)", "Country", os.path.join(results_dir, f"{safe_filename(base_name)}_Top_30_Countries_by_Open_Access_Percentage.jpeg"), n=30, sort_value_col="Open Access Percentage", value_format="float")
        difference_metric = "Unadjusted Open Access Minus Not Open Access Mean Annualized Citation Rate"
        difference_plot = country_open_access.dropna(subset=[difference_metric]).copy()
        difference_plot = difference_plot[difference_plot["Open Access Publications"] >= 5]
        difference_plot = difference_plot[difference_plot["Not Open Access Publications"] >= 5]
        save_horizontal_bar_from_table(difference_plot, difference_metric, "Country", f"Countries by Unadjusted Open Access Minus Not Open Access Mean Annualized Citation Rate - {condition}", "Difference in Mean Annualized Citation Rate", "Country", os.path.join(results_dir, f"{safe_filename(base_name)}_Countries_by_Unadjusted_Open_Access_Minus_Not_Open_Access_Mean_Annualized_Citation_Rate.jpeg"), n=30, sort_value_col=difference_metric, value_format="float")
    if not annual_open_access.empty:
        save_year_bar(annual_open_access["Year"].values, annual_open_access["Open Access Percentage"].values, f"Open Access Percentage by Publication Year - {condition}", "Open Access Papers (%)", os.path.join(results_dir, f"{safe_filename(base_name)}_Open_Access_Percentage_by_Publication_Year.jpeg"), value_format="float")
    return status_summary, unadjusted_comparison, country_open_access, annual_open_access

def annual_publication_and_citation_impact_tables(df):
    valid = df.dropna(subset=["Year"]).copy()
    if valid.empty:
        return pd.Series(dtype=int), pd.Series(dtype=int)
    valid["Year"] = valid["Year"].astype(int)
    publications = valid["Year"].value_counts().sort_index()
    citation_impact = valid.groupby("Year")["Cited by"].sum().sort_index()
    return publications, citation_impact


def add_annualized_citation_metrics(df, citation_reference_year=None):
    output = df.copy()
    output["Year"] = pd.to_numeric(output["Year"], errors="coerce")
    valid_years = output["Year"].dropna()
    if citation_reference_year is None:
        citation_reference_year = int(valid_years.max()) if not valid_years.empty else None
    elif not valid_years.empty and int(citation_reference_year) < int(valid_years.max()):
        raise ValueError("The citation reference year cannot be earlier than the latest publication year in the dataset.")
    if citation_reference_year is None:
        output["Citation Years"] = np.nan
        output["Citations per Year"] = np.nan
        return output, citation_reference_year
    citation_reference_year = int(citation_reference_year)
    citation_years = citation_reference_year - output["Year"] + 1
    citation_years = citation_years.where(citation_years > 0, np.nan)
    output["Citation Years"] = citation_years
    output["Citations per Year"] = output["Cited by"] / output["Citation Years"]
    output.loc[output["Year"].isna(), ["Citation Years", "Citations per Year"]] = np.nan
    return output, citation_reference_year

def annualized_citation_rate_tables(df):
    valid = df.dropna(subset=["Year", "Citations per Year"]).copy()
    if valid.empty:
        return pd.Series(dtype=float), pd.Series(dtype=float)
    valid["Year"] = valid["Year"].astype(int)
    total_annualized_rate = valid.groupby("Year")["Citations per Year"].sum().sort_index()
    mean_annualized_rate = valid.groupby("Year")["Citations per Year"].mean().sort_index()
    return total_annualized_rate, mean_annualized_rate

def create_collaboration_graph(entity_lists, allowed_entities=None):
    allowed_set = set(allowed_entities) if allowed_entities is not None else None
    graph = nx.Graph()
    if allowed_set is not None:
        graph.add_nodes_from(sorted(allowed_set))
    for entities in entity_lists:
        full_entities = list(dict.fromkeys([entity for entity in entities if clean_text(entity)]))
        if not full_entities:
            continue
        selected_entities = [entity for entity in full_entities if allowed_set is None or entity in allowed_set]
        graph.add_nodes_from(selected_entities)
        if len(selected_entities) < 2:
            continue
        fractional_contribution = 1.0 / max(len(full_entities) - 1, 1)
        for source, target in combinations(sorted(selected_entities), 2):
            if graph.has_edge(source, target):
                graph[source][target]["shared_publications"] += 1
                graph[source][target]["fractional_weight"] += fractional_contribution
            else:
                graph.add_edge(
                    source,
                    target,
                    shared_publications=1,
                    fractional_weight=fractional_contribution
                )
    graph.remove_edges_from(nx.selfloop_edges(graph))
    for source, target, data in graph.edges(data=True):
        fractional_weight = float(data.get("fractional_weight", 0))
        data["weight"] = fractional_weight
        data["distance"] = 1.0 / fractional_weight if fractional_weight > 0 else 1.0
        data["shared_publications"] = int(data.get("shared_publications", 0))
    return graph


def detect_collaboration_communities(graph):
    if graph.number_of_nodes() == 0:
        return [], {}, {}, "", 0
    if graph.number_of_edges() == 0:
        communities = [{node} for node in sorted(graph.nodes())]
        algorithm = "singleton_components"
        modularity = 0.0
    else:
        algorithm = "louvain"
        try:
            communities = nx.algorithms.community.louvain_communities(
                graph,
                weight="fractional_weight",
                seed=42
            )
        except Exception:
            algorithm = "greedy_modularity"
            communities = nx.algorithms.community.greedy_modularity_communities(
                graph,
                weight="fractional_weight"
            )
        communities = [set(community) for community in communities if community]
        try:
            modularity = nx.algorithms.community.modularity(
                graph,
                communities,
                weight="fractional_weight"
            )
        except Exception:
            modularity = 0.0
    communities = sorted(
        communities,
        key=lambda community: (
            -len(community),
            -sum(
                float(graph[source][target].get("fractional_weight", 0))
                for source, target in graph.subgraph(community).edges()
            ),
            sorted(community)[0]
        )
    )
    community_map = {}
    community_sizes = {}
    for index, community in enumerate(communities, 1):
        for node in community:
            community_map[node] = index
            community_sizes[node] = len(community)
    return communities, community_map, community_sizes, algorithm, float(modularity)


def component_membership(graph):
    components = sorted(
        nx.connected_components(graph),
        key=lambda component: (-len(component), sorted(component)[0])
    )
    component_map = {}
    component_sizes = {}
    for index, component in enumerate(components, 1):
        for node in component:
            component_map[node] = index
            component_sizes[node] = len(component)
    return components, component_map, component_sizes


def select_visualization_backbone(graph, edges_per_node):
    if graph.number_of_edges() == 0:
        return set()
    if graph.number_of_edges() <= max(220, graph.number_of_nodes() * edges_per_node):
        return {tuple(sorted((source, target))) for source, target in graph.edges()}
    selected_edges = set()
    for component in nx.connected_components(graph):
        subgraph = graph.subgraph(component)
        if subgraph.number_of_edges() == 0:
            continue
        tree = nx.maximum_spanning_tree(subgraph, weight="fractional_weight")
        selected_edges.update(tuple(sorted((source, target))) for source, target in tree.edges())
    for node in graph.nodes():
        incident_edges = sorted(
            graph.edges(node, data=True),
            key=lambda edge: (
                -float(edge[2].get("fractional_weight", 0)),
                -int(edge[2].get("shared_publications", 0)),
                str(edge[1]).casefold()
            )
        )
        for source, target, data in incident_edges[:edges_per_node]:
            selected_edges.add(tuple(sorted((source, target))))
    return selected_edges


def collaboration_network_layout(graph, community_map, seed=42):
    positions = {}
    active_nodes = [node for node in graph.nodes() if graph.degree(node) > 0]
    isolate_nodes = sorted([node for node in graph.nodes() if graph.degree(node) == 0])
    if active_nodes:
        active_graph = graph.subgraph(active_nodes).copy()
        active_communities = sorted(set(community_map[node] for node in active_nodes))
        community_graph = nx.Graph()
        community_graph.add_nodes_from(active_communities)
        for source, target, data in active_graph.edges(data=True):
            source_community = community_map[source]
            target_community = community_map[target]
            if source_community == target_community:
                continue
            weight = float(data.get("fractional_weight", 0))
            if community_graph.has_edge(source_community, target_community):
                community_graph[source_community][target_community]["weight"] += weight
            else:
                community_graph.add_edge(source_community, target_community, weight=weight)
        if len(active_communities) == 1:
            community_centers = {active_communities[0]: np.array([0.0, 0.0])}
        elif community_graph.number_of_edges() == 0:
            community_centers = nx.circular_layout(community_graph, scale=4.5)
        else:
            community_centers = nx.spring_layout(
                community_graph,
                seed=seed,
                weight="weight",
                k=max(1.4, 3.0 / np.sqrt(len(active_communities))),
                iterations=500,
                scale=4.5
            )
        for community_id in active_communities:
            community_nodes = sorted(
                [node for node in active_nodes if community_map[node] == community_id]
            )
            center = np.asarray(community_centers[community_id], dtype=float)
            if len(community_nodes) == 1:
                positions[community_nodes[0]] = center
                continue
            subgraph = active_graph.subgraph(community_nodes).copy()
            local_positions = nx.spring_layout(
                subgraph,
                seed=seed + int(community_id),
                weight="fractional_weight",
                k=max(0.35, 1.6 / np.sqrt(len(community_nodes))),
                iterations=500,
                scale=1.0
            )
            local_radius = 0.65 + 0.11 * np.sqrt(len(community_nodes))
            for node in community_nodes:
                positions[node] = center + np.asarray(local_positions[node]) * local_radius
    if isolate_nodes:
        if positions:
            active_radius = max(float(np.linalg.norm(position)) for position in positions.values())
        else:
            active_radius = 1.0
        isolate_radius = max(5.8, active_radius + 1.8)
        angles = np.linspace(0, 2 * np.pi, len(isolate_nodes), endpoint=False)
        for node, angle in zip(isolate_nodes, angles):
            positions[node] = np.array([
                isolate_radius * np.cos(angle),
                isolate_radius * np.sin(angle)
            ])
    return positions


def network_graph_statistics(graph, communities, modularity, community_algorithm, displayed_edges):
    components = sorted(
        nx.connected_components(graph),
        key=lambda component: (-len(component), sorted(component)[0])
    ) if graph.number_of_nodes() else []
    largest_component = components[0] if components else set()
    largest_component_size = len(largest_component)
    largest_component_share = largest_component_size / graph.number_of_nodes() * 100 if graph.number_of_nodes() else 0
    if largest_component_size > 1:
        largest_subgraph = graph.subgraph(largest_component).copy()
        average_path_length = nx.average_shortest_path_length(
            largest_subgraph,
            weight="distance"
        )
        path_lengths = dict(nx.all_pairs_dijkstra_path_length(largest_subgraph, weight="distance"))
        weighted_diameter = max(
            max(lengths.values()) for lengths in path_lengths.values() if lengths
        )
    else:
        average_path_length = 0.0
        weighted_diameter = 0.0
    try:
        assortativity = float(nx.degree_assortativity_coefficient(graph)) if graph.number_of_edges() else 0.0
        if not np.isfinite(assortativity):
            assortativity = 0.0
    except Exception:
        assortativity = 0.0
    average_clustering = float(nx.average_clustering(graph, weight="fractional_weight")) if graph.number_of_nodes() else 0.0
    return {
        "Network Nodes": int(graph.number_of_nodes()),
        "Network Edges": int(graph.number_of_edges()),
        "Displayed Visualization Edges": int(len(displayed_edges)),
        "Isolates": int(nx.number_of_isolates(graph)),
        "Network Density": float(nx.density(graph)) if graph.number_of_nodes() > 1 else 0.0,
        "Connected Components": int(len(components)),
        "Largest Component Size": int(largest_component_size),
        "Largest Component Share (%)": float(largest_component_share),
        "Communities": int(len(communities)),
        "Non-Singleton Communities": int(sum(1 for community in communities if len(community) > 1)),
        "Community Detection Algorithm": community_algorithm,
        "Network Modularity": float(modularity),
        "Average Clustering Coefficient": average_clustering,
        "Transitivity": float(nx.transitivity(graph)) if graph.number_of_nodes() else 0.0,
        "Degree Assortativity": assortativity,
        "Average Weighted Shortest Path in Largest Component": float(average_path_length),
        "Weighted Diameter of Largest Component": float(weighted_diameter)
    }


def save_collaboration_network_figure(graph, ranked_nodes, community_map, displayed_edges, summary, title, selection_text, output_path):
    positions = collaboration_network_layout(graph, community_map, seed=42)
    figure = plt.figure(figsize=(36, 24))
    grid = figure.add_gridspec(1, 2, width_ratios=[2.25, 1.15], wspace=0.02)
    network_axis = figure.add_subplot(grid[0, 0])
    key_axis = figure.add_subplot(grid[0, 1])
    community_ids = sorted(set(community_map.values()))
    color_map = plt.colormaps.get_cmap("tab20").resampled(max(len(community_ids), 1))
    community_colors = {
        community_id: color_map(index)
        for index, community_id in enumerate(community_ids)
    }
    collaboration_strength = dict(graph.degree(weight="shared_publications"))
    maximum_strength = max(collaboration_strength.values()) if collaboration_strength else 1
    maximum_strength = max(maximum_strength, 1)
    node_sizes = [
        260 + 1500 * np.log1p(collaboration_strength.get(node, 0)) / np.log1p(maximum_strength)
        for node in graph.nodes()
    ]
    node_colors = [
        community_colors.get(community_map.get(node), (0.7, 0.7, 0.7, 1.0))
        if graph.degree(node) > 0 else (0.75, 0.75, 0.75, 1.0)
        for node in graph.nodes()
    ]
    intra_edges = []
    inter_edges = []
    intra_widths = []
    inter_widths = []
    displayed_weights = [
        int(graph[source][target].get("shared_publications", 1))
        for source, target in displayed_edges
    ]
    maximum_edge_weight = max(displayed_weights) if displayed_weights else 1
    for source, target in displayed_edges:
        width = 0.35 + 4.2 * np.log1p(graph[source][target].get("shared_publications", 1)) / np.log1p(maximum_edge_weight)
        if community_map.get(source) == community_map.get(target):
            intra_edges.append((source, target))
            intra_widths.append(width)
        else:
            inter_edges.append((source, target))
            inter_widths.append(width)
    if inter_edges:
        nx.draw_networkx_edges(
            graph,
            positions,
            edgelist=inter_edges,
            width=inter_widths,
            alpha=0.22,
            edge_color="gray",
            ax=network_axis
        )
    if intra_edges:
        nx.draw_networkx_edges(
            graph,
            positions,
            edgelist=intra_edges,
            width=intra_widths,
            alpha=0.45,
            edge_color="dimgray",
            ax=network_axis
        )
    nx.draw_networkx_nodes(
        graph,
        positions,
        node_size=node_sizes,
        node_color=node_colors,
        edgecolors="white",
        linewidths=1.0,
        alpha=0.95,
        ax=network_axis
    )
    collaboration_rank = {node: rank for rank, node in enumerate(ranked_nodes, 1)}
    rank_labels = {node: str(collaboration_rank[node]) for node in graph.nodes()}
    nx.draw_networkx_labels(
        graph,
        positions,
        labels=rank_labels,
        font_size=7,
        font_color="white",
        font_weight="bold",
        font_family="sans-serif",
        ax=network_axis
    )
    network_axis.set_axis_off()
    network_axis.set_title(
        title,
        fontsize=25,
        fontweight="bold",
        pad=22
    )
    network_axis.text(
        0.5,
        1.005,
        selection_text,
        transform=network_axis.transAxes,
        ha="center",
        va="bottom",
        fontsize=12
    )
    network_axis.text(
        0.01,
        0.01,
        "Node number = collaboration rank | Node size = total link strength | Color = Louvain community | Edge width = shared publications | Display uses a strongest-link backbone; all metrics use the complete induced network",
        transform=network_axis.transAxes,
        ha="left",
        va="bottom",
        fontsize=10,
        bbox={"boxstyle": "round,pad=0.45", "facecolor": "white", "alpha": 0.92, "edgecolor": "lightgray"}
    )
    key_axis.set_axis_off()
    key_axis.set_xlim(0, 1)
    key_axis.set_ylim(0, 1)
    key_axis.text(
        0.02,
        0.985,
        "Collaboration rank",
        fontsize=18,
        fontweight="bold",
        va="top"
    )
    key_axis.text(
        0.02,
        0.955,
        "Ranked by total link strength within this network",
        fontsize=10,
        va="top"
    )
    number_of_nodes = len(ranked_nodes)
    key_columns = 2 if number_of_nodes > 52 else 1
    rows_per_column = int(np.ceil(number_of_nodes / key_columns)) if number_of_nodes else 1
    top_y = 0.925
    bottom_y = 0.18
    line_spacing = (top_y - bottom_y) / max(rows_per_column - 1, 1)
    font_size = 7.2 if rows_per_column > 42 else 8.2
    for index, node in enumerate(ranked_nodes):
        column = index // rows_per_column
        row = index % rows_per_column
        x_position = 0.02 + column * 0.50
        y_position = top_y - row * line_spacing
        community_id = community_map.get(node, 0)
        marker_color = community_colors.get(community_id, (0.7, 0.7, 0.7, 1.0))
        if graph.degree(node) == 0:
            marker_color = (0.75, 0.75, 0.75, 1.0)
        key_axis.scatter(
            [x_position + 0.008],
            [y_position],
            s=22,
            color=[marker_color],
            edgecolors="none"
        )
        key_axis.text(
            x_position + 0.022,
            y_position,
            f"{index + 1}. {node} (C{community_id}; TLS {int(round(collaboration_strength.get(node, 0)))})",
            fontsize=font_size,
            va="center",
            ha="left"
        )
    metric_lines = [
        f"Nodes: {summary['Network Nodes']}    Edges: {summary['Network Edges']}    Isolates: {summary['Isolates']}",
        f"Density: {summary['Network Density']:.3f}    Components: {summary['Connected Components']}",
        f"Communities: {summary['Communities']}    Modularity: {summary['Network Modularity']:.3f}",
        f"Average clustering: {summary['Average Clustering Coefficient']:.3f}    Transitivity: {summary['Transitivity']:.3f}",
        f"Largest component: {summary['Largest Component Size']} nodes ({summary['Largest Component Share (%)']:.1f}%)",
        f"Displayed edges: {summary['Displayed Visualization Edges']} of {summary['Network Edges']}"
    ]
    key_axis.text(
        0.02,
        0.135,
        "Network structure",
        fontsize=14,
        fontweight="bold",
        va="top"
    )
    key_axis.text(
        0.02,
        0.112,
        "\n".join(metric_lines),
        fontsize=9.5,
        va="top",
        linespacing=1.35
    )
    figure.savefig(output_path, dpi=300, format="jpeg", bbox_inches="tight")
    plt.close(figure)


def save_ranked_collaboration_network(entity_lists, entity_table, entity_column, selected_entities, selection_ranks, selection_criterion, network_title, base_name, output_folder, edges_per_node):
    selected_entities = list(dict.fromkeys([entity for entity in selected_entities if clean_text(entity)]))
    if not selected_entities:
        return None
    graph = create_collaboration_graph(entity_lists, allowed_entities=selected_entities)
    metric_table = entity_table.set_index(entity_column, drop=False)
    for node in graph.nodes():
        if node in metric_table.index:
            row = metric_table.loc[node]
            if isinstance(row, pd.DataFrame):
                row = row.iloc[0]
            graph.nodes[node]["publications"] = int(row.get("Publications", 0))
            graph.nodes[node]["citations"] = int(row.get("Citations", 0))
            graph.nodes[node]["normalized_citation_impact"] = float(row.get("Total Annualized Citation Rate", 0))
            graph.nodes[node]["average_citations_per_year_per_paper"] = float(row.get("Mean Annualized Citation Rate per Paper", 0))
        else:
            graph.nodes[node]["publications"] = 0
            graph.nodes[node]["citations"] = 0
            graph.nodes[node]["normalized_citation_impact"] = 0.0
            graph.nodes[node]["average_citations_per_year_per_paper"] = 0.0
        graph.nodes[node]["selection_rank"] = int(selection_ranks.get(node, 0))
        graph.nodes[node]["selection_criterion"] = selection_criterion
    distinct_collaborators = dict(graph.degree())
    total_link_strength = dict(graph.degree(weight="shared_publications"))
    fractional_link_strength = dict(graph.degree(weight="fractional_weight"))
    degree_centrality = nx.degree_centrality(graph) if graph.number_of_nodes() > 1 else {node: 0.0 for node in graph.nodes()}
    betweenness_centrality = nx.betweenness_centrality(
        graph,
        weight="distance",
        normalized=True
    ) if graph.number_of_nodes() else {}
    closeness_centrality = nx.closeness_centrality(
        graph,
        distance="distance"
    ) if graph.number_of_nodes() else {}
    clustering_coefficients = nx.clustering(
        graph,
        weight="fractional_weight"
    ) if graph.number_of_nodes() else {}
    pagerank = nx.pagerank(
        graph,
        weight="fractional_weight"
    ) if graph.number_of_nodes() else {}
    communities, community_map, community_sizes, community_algorithm, modularity = detect_collaboration_communities(graph)
    components, component_map, component_sizes = component_membership(graph)
    displayed_edges = select_visualization_backbone(graph, edges_per_node=edges_per_node)
    summary = network_graph_statistics(
        graph,
        communities,
        modularity,
        community_algorithm,
        displayed_edges
    )
    summary.update({
        "Network": network_title,
        "Selection Criterion": selection_criterion,
        "Selected Entities": len(selected_entities)
    })
    node_metrics = {}
    for node in graph.nodes():
        node_metrics[node] = {
            "distinct_collaborators": int(distinct_collaborators.get(node, 0)),
            "total_link_strength": float(total_link_strength.get(node, 0)),
            "fractional_link_strength": float(fractional_link_strength.get(node, 0)),
            "degree_centrality": float(degree_centrality.get(node, 0)),
            "betweenness_centrality": float(betweenness_centrality.get(node, 0)),
            "closeness_centrality": float(closeness_centrality.get(node, 0)),
            "clustering_coefficient": float(clustering_coefficients.get(node, 0)),
            "pagerank": float(pagerank.get(node, 0))
        }
    ranked_nodes = sorted(
        graph.nodes(),
        key=lambda node: (
            -node_metrics[node]["total_link_strength"],
            -node_metrics[node]["distinct_collaborators"],
            -node_metrics[node]["betweenness_centrality"],
            node.casefold()
        )
    )
    for collaboration_rank, node in enumerate(ranked_nodes, 1):
        graph.nodes[node]["collaboration_rank"] = collaboration_rank
        graph.nodes[node]["distinct_collaborators"] = node_metrics[node]["distinct_collaborators"]
        graph.nodes[node]["total_link_strength"] = node_metrics[node]["total_link_strength"]
        graph.nodes[node]["fractional_link_strength"] = node_metrics[node]["fractional_link_strength"]
        graph.nodes[node]["degree_centrality"] = node_metrics[node]["degree_centrality"]
        graph.nodes[node]["betweenness_centrality"] = node_metrics[node]["betweenness_centrality"]
        graph.nodes[node]["closeness_centrality"] = node_metrics[node]["closeness_centrality"]
        graph.nodes[node]["clustering_coefficient"] = node_metrics[node]["clustering_coefficient"]
        graph.nodes[node]["pagerank"] = node_metrics[node]["pagerank"]
        graph.nodes[node]["community"] = int(community_map.get(node, 0))
        graph.nodes[node]["community_size"] = int(community_sizes.get(node, 0))
        graph.nodes[node]["component"] = int(component_map.get(node, 0))
        graph.nodes[node]["component_size"] = int(component_sizes.get(node, 0))
    graph.graph.update({key: value for key, value in summary.items() if isinstance(value, (str, int, float))})
    ranking_rows = []
    for collaboration_rank, node in enumerate(ranked_nodes, 1):
        ranking_rows.append({
            "Collaboration Rank": collaboration_rank,
            "Name": node,
            "Selection Criterion": selection_criterion,
            "Selection Rank": selection_ranks.get(node, ""),
            "Publications": graph.nodes[node]["publications"],
            "Citations": graph.nodes[node]["citations"],
            "Total Annualized Citation Rate": graph.nodes[node]["normalized_citation_impact"],
            "Mean Annualized Citation Rate per Paper": graph.nodes[node]["average_citations_per_year_per_paper"],
            "Distinct Collaborators": node_metrics[node]["distinct_collaborators"],
            "Total Link Strength (Shared Publications)": node_metrics[node]["total_link_strength"],
            "Fractional Collaboration Strength": node_metrics[node]["fractional_link_strength"],
            "Degree Centrality": node_metrics[node]["degree_centrality"],
            "Betweenness Centrality": node_metrics[node]["betweenness_centrality"],
            "Closeness Centrality": node_metrics[node]["closeness_centrality"],
            "Clustering Coefficient": node_metrics[node]["clustering_coefficient"],
            "PageRank": node_metrics[node]["pagerank"],
            "Community": community_map.get(node, 0),
            "Community Size": community_sizes.get(node, 0),
            "Component": component_map.get(node, 0),
            "Component Size": component_sizes.get(node, 0)
        })
    ranking_path = os.path.join(output_folder, f"{base_name}_Ranking.csv")
    pd.DataFrame(ranking_rows).to_csv(ranking_path, index=False, encoding="utf-8-sig")
    save_collaboration_network_figure(
        graph,
        ranked_nodes,
        community_map,
        displayed_edges,
        summary,
        network_title,
        selection_criterion,
        os.path.join(output_folder, f"{base_name}_Network.jpeg")
    )
    return summary


def save_combined_network_analysis(combined_df, output_dir, citation_reference_year=None):
    network_dir = os.path.join(output_dir, "optimized_network_analysis")
    os.makedirs(network_dir, exist_ok=True)
    obsolete_suffixes = ("_Edges.csv", "_Communities.csv", "_Network_Summary.csv", "_Network.graphml")
    for filename in os.listdir(network_dir):
        if filename.startswith("Combined_PID_IEI_") and filename.endswith(obsolete_suffixes):
            os.remove(os.path.join(network_dir, filename))
    combined_df, citation_reference_year = add_annualized_citation_metrics(combined_df, citation_reference_year)
    author_lists = entities_per_row(combined_df, "Author")
    country_lists = entities_per_row(combined_df, "Country")
    author_table = productivity_and_citations(combined_df, author_lists, "Author")
    country_table = productivity_and_citations(combined_df, country_lists, "Country")
    top_cited_authors = author_table.sort_values(
        ["Citations", "Publications", "Author"],
        ascending=[False, False, True]
    ).head(100)
    top_productive_authors = author_table.sort_values(
        ["Publications", "Citations", "Author"],
        ascending=[False, False, True]
    ).head(100)
    citation_selection_ranks = {
        author: rank
        for rank, author in enumerate(top_cited_authors["Author"].tolist(), 1)
    }
    productivity_selection_ranks = {
        author: rank
        for rank, author in enumerate(top_productive_authors["Author"].tolist(), 1)
    }
    country_selection_ranks = {
        country: rank
        for rank, country in enumerate(
            country_table.sort_values(
                ["Publications", "Citations", "Country"],
                ascending=[False, False, True]
            )["Country"].tolist(),
            1
        )
    }
    summaries = []
    cited_summary = save_ranked_collaboration_network(
        author_lists,
        author_table,
        "Author",
        top_cited_authors["Author"].tolist(),
        citation_selection_ranks,
        "Top 100 authors by total citations in the combined deduplicated PID + IEI corpus",
        "Combined PID + IEI — Top 100 Authors by Citations: Collaboration Network",
        "Combined_PID_IEI_Top_100_Authors_by_Citations_Collaboration",
        network_dir,
        edges_per_node=3
    )
    if cited_summary:
        summaries.append(cited_summary)
    productive_summary = save_ranked_collaboration_network(
        author_lists,
        author_table,
        "Author",
        top_productive_authors["Author"].tolist(),
        productivity_selection_ranks,
        "Top 100 authors by publication count in the combined deduplicated PID + IEI corpus",
        "Combined PID + IEI — Top 100 Authors by Publications: Collaboration Network",
        "Combined_PID_IEI_Top_100_Authors_by_Publications_Collaboration",
        network_dir,
        edges_per_node=3
    )
    if productive_summary:
        summaries.append(productive_summary)
    country_summary = save_ranked_collaboration_network(
        country_lists,
        country_table,
        "Country",
        country_table["Country"].tolist(),
        country_selection_ranks,
        "All countries represented in the combined deduplicated PID + IEI corpus",
        "Combined PID + IEI — All Countries: International Collaboration Network",
        "Combined_PID_IEI_All_Countries_Collaboration",
        network_dir,
        edges_per_node=4
    )
    if country_summary:
        summaries.append(country_summary)
    if summaries:
        summary_table = pd.DataFrame(summaries)
        summary_table.insert(0, "Citation Annualization Reference Year", citation_reference_year if citation_reference_year is not None else "")
        summary_table.to_csv(
            os.path.join(network_dir, "Combined_PID_IEI_Network_Analysis_Summary.csv"),
            index=False,
            encoding="utf-8-sig"
        )


def dataset_label(dataset_name):
    if dataset_name == "primary_immunodeficiency_diseases":
        return "Primary Immunodeficiency Diseases"
    if dataset_name == "inborn_errors_of_immunity":
        return "Inborn Errors of Immunity"
    return dataset_name.replace("_", " ").title()


def analyze_dataset(df, dataset_name, output_dir, citation_reference_year=None):
    condition = dataset_label(dataset_name)
    base_name = condition
    safe_base_name = safe_filename(base_name)
    results_dir = os.path.join(output_dir, "bibliometric_results")
    network_dir = os.path.join(output_dir, "optimized_network_analysis")
    tables_dir = os.path.join(output_dir, "bibliometric_tables")
    os.makedirs(results_dir, exist_ok=True)
    os.makedirs(network_dir, exist_ok=True)
    os.makedirs(tables_dir, exist_ok=True)
    remove_existing_files([
        os.path.join(tables_dir, f"{safe_base_name}_authors_normalized_citation_impact.csv"),
        os.path.join(tables_dir, f"{safe_base_name}_countries_normalized_citation_impact.csv"),
        os.path.join(tables_dir, f"{safe_base_name}_journals_normalized_citation_impact.csv"),
        os.path.join(tables_dir, f"{safe_base_name}_publishers_normalized_citation_impact.csv"),
        os.path.join(tables_dir, f"{safe_base_name}_cumulative_citation_impact_by_publication_year.csv"),
        os.path.join(tables_dir, f"{safe_base_name}_normalized_citation_impact_by_publication_year.csv"),
        os.path.join(tables_dir, f"{safe_base_name}_open_access_citation_advantage_summary.csv"),
        os.path.join(results_dir, f"{safe_base_name}_Top_30_Authors_by_Normalized_Citation_Impact.jpeg"),
        os.path.join(results_dir, f"{safe_base_name}_Top_30_Countries_by_Normalized_Citation_Impact.jpeg"),
        os.path.join(results_dir, f"{safe_base_name}_Top_30_Journals_by_Normalized_Citation_Impact.jpeg"),
        os.path.join(results_dir, f"{safe_base_name}_Top_10_Publishers_by_Normalized_Citation_Impact.jpeg"),
        os.path.join(results_dir, f"{safe_base_name}_Cumulative_Citation_Impact_by_Publication_Year.jpeg"),
        os.path.join(results_dir, f"{safe_base_name}_Normalized_Citation_Impact_by_Publication_Year.jpeg"),
        os.path.join(results_dir, f"{safe_base_name}_Average_Citations_per_Year_per_Paper_by_Publication_Year.jpeg"),
        os.path.join(results_dir, f"{safe_base_name}_Open_Access_Average_Citations_per_Year_per_Paper.jpeg"),
        os.path.join(results_dir, f"{safe_base_name}_Top_Countries_by_Open_Access_Normalized_Citation_Advantage.jpeg")
    ])
    df, citation_reference_year = add_annualized_citation_metrics(df, citation_reference_year)
    annualized_metric = "Total Annualized Citation Rate"
    author_lists = entities_per_row(df, "Author")
    country_lists = entities_per_row(df, "Country")
    authors = flatten_unique(author_lists)
    countries = flatten_unique(country_lists)
    author_table = productivity_and_citations(df, author_lists, "Author")
    country_table = productivity_and_citations(df, country_lists, "Country")
    journal_table = single_column_productivity_and_citations(df, "Source title", "Journal")
    publisher_table = single_column_productivity_and_citations(df, "Publisher", "Publisher")
    author_table.to_csv(os.path.join(tables_dir, f"{safe_base_name}_authors_productivity_impact.csv"), index=False, encoding="utf-8-sig")
    country_table.to_csv(os.path.join(tables_dir, f"{safe_base_name}_countries_productivity_impact.csv"), index=False, encoding="utf-8-sig")
    journal_table.to_csv(os.path.join(tables_dir, f"{safe_base_name}_journals_productivity_impact.csv"), index=False, encoding="utf-8-sig")
    publisher_table.to_csv(os.path.join(tables_dir, f"{safe_base_name}_publishers_productivity_impact.csv"), index=False, encoding="utf-8-sig")
    save_horizontal_bar_from_series(authors, f"Top 30 Authors by Publication Count - {condition}", "Number of Papers", "Author", os.path.join(results_dir, f"{safe_base_name}_Top_30_Authors_by_Publication_Count.jpeg"), n=30)
    save_horizontal_bar_from_table(author_table.sort_values(["Citations", "Publications", "Author"], ascending=[False, False, True]), "Citations", "Author", f"Top 30 Authors by Citation Count - {condition}", "Total Citations", "Author", os.path.join(results_dir, f"{safe_base_name}_Top_30_Authors_by_Citation_Count.jpeg"), n=30, sort_value_col="Citations")
    save_horizontal_bar_from_table(author_table.sort_values([annualized_metric, "Publications", "Author"], ascending=[False, False, True]), annualized_metric, "Author", f"Top 30 Authors by Total Annualized Citation Rate - {condition}", "Total Annualized Citation Rate", "Author", os.path.join(results_dir, f"{safe_base_name}_Top_30_Authors_by_Total_Annualized_Citation_Rate.jpeg"), n=30, sort_value_col=annualized_metric, value_format="float")
    save_horizontal_bar_from_series(countries, f"Top 30 Countries by Publication Count - {condition}", "Number of Papers", "Country", os.path.join(results_dir, f"{safe_base_name}_Top_30_Countries_by_Publication_Count.jpeg"), n=30)
    save_horizontal_bar_from_table(country_table.sort_values(["Citations", "Publications", "Country"], ascending=[False, False, True]), "Citations", "Country", f"Top 30 Countries by Citation Count - {condition}", "Total Citations", "Country", os.path.join(results_dir, f"{safe_base_name}_Top_30_Countries_by_Citation_Count.jpeg"), n=30, sort_value_col="Citations")
    save_horizontal_bar_from_table(country_table.sort_values([annualized_metric, "Publications", "Country"], ascending=[False, False, True]), annualized_metric, "Country", f"Top 30 Countries by Total Annualized Citation Rate - {condition}", "Total Annualized Citation Rate", "Country", os.path.join(results_dir, f"{safe_base_name}_Top_30_Countries_by_Total_Annualized_Citation_Rate.jpeg"), n=30, sort_value_col=annualized_metric, value_format="float")
    save_vertical_count_bar(df["Language of Original Document"].astype(str).str.strip(), f"Top 10 Languages - {condition}", "Language", "Number of Papers", os.path.join(results_dir, f"{safe_base_name}_Top_10_Languages.jpeg"), n=10)
    save_open_access_pie(df, f"Open Access Status - {condition}", os.path.join(results_dir, f"{safe_base_name}_Open_Access_Status.jpeg"))
    open_access_status_summary, open_access_unadjusted_comparison, open_access_country_summary, open_access_annual_summary = save_open_access_extended_analysis(df, country_lists, condition, base_name, results_dir, tables_dir)
    save_horizontal_bar_from_series(df["Source title"].astype(str).str.strip(), f"Top 30 Journals by Publication Count - {condition}", "Number of Papers", "Journal", os.path.join(results_dir, f"{safe_base_name}_Top_30_Journals_by_Publication_Count.jpeg"), n=30)
    save_horizontal_bar_from_table(journal_table.sort_values(["Citations", "Publications", "Journal"], ascending=[False, False, True]), "Citations", "Journal", f"Top 30 Journals by Citation Count - {condition}", "Total Citations", "Journal", os.path.join(results_dir, f"{safe_base_name}_Top_30_Journals_by_Citation_Count.jpeg"), n=30, sort_value_col="Citations")
    save_horizontal_bar_from_table(journal_table.sort_values([annualized_metric, "Publications", "Journal"], ascending=[False, False, True]), annualized_metric, "Journal", f"Top 30 Journals by Total Annualized Citation Rate - {condition}", "Total Annualized Citation Rate", "Journal", os.path.join(results_dir, f"{safe_base_name}_Top_30_Journals_by_Total_Annualized_Citation_Rate.jpeg"), n=30, sort_value_col=annualized_metric, value_format="float")
    save_horizontal_bar_from_series(df["Publisher"].astype(str).str.strip(), f"Top 10 Publishers by Publication Count - {condition}", "Number of Papers", "Publisher", os.path.join(results_dir, f"{safe_base_name}_Top_10_Publishers_by_Publication_Count.jpeg"), n=10)
    save_horizontal_bar_from_table(publisher_table.sort_values(["Citations", "Publications", "Publisher"], ascending=[False, False, True]), "Citations", "Publisher", f"Top 10 Publishers by Citation Count - {condition}", "Total Citations", "Publisher", os.path.join(results_dir, f"{safe_base_name}_Top_10_Publishers_by_Citation_Count.jpeg"), n=10, sort_value_col="Citations")
    save_horizontal_bar_from_table(publisher_table.sort_values([annualized_metric, "Publications", "Publisher"], ascending=[False, False, True]), annualized_metric, "Publisher", f"Top 10 Publishers by Total Annualized Citation Rate - {condition}", "Total Annualized Citation Rate", "Publisher", os.path.join(results_dir, f"{safe_base_name}_Top_10_Publishers_by_Total_Annualized_Citation_Rate.jpeg"), n=10, sort_value_col=annualized_metric, value_format="float")
    year_counts, year_citation_impact = annual_publication_and_citation_impact_tables(df)
    year_annualized_impact, year_mean_annualized_impact = annualized_citation_rate_tables(df)
    year_counts_for_descriptive_statistics = year_counts.reindex(range(int(year_counts.index.min()), citation_reference_year + 1), fill_value=0) if not year_counts.empty and citation_reference_year is not None else year_counts
    descriptive_statistics = descriptive_statistics_table(df, condition, author_lists, country_lists, author_table, country_table, journal_table, publisher_table, year_counts_for_descriptive_statistics, year_citation_impact, year_annualized_impact, year_mean_annualized_impact)
    if not year_counts.empty:
        pd.DataFrame({"Year": year_counts.index.astype(int), "Publications": year_counts.values}).to_csv(os.path.join(tables_dir, f"{safe_base_name}_publications_per_year.csv"), index=False, encoding="utf-8-sig")
        save_year_bar(year_counts.index, year_counts.values, f"Publications per Year - {condition}", "Number of Papers", os.path.join(results_dir, f"{safe_base_name}_Publications_per_Year.jpeg"))
    if not year_citation_impact.empty:
        pd.DataFrame({"Publication Year": year_citation_impact.index.astype(int), "Total Citations by Publication-Year Cohort": year_citation_impact.values}).to_csv(os.path.join(tables_dir, f"{safe_base_name}_total_citations_by_publication_year_cohort.csv"), index=False, encoding="utf-8-sig")
        save_year_bar(year_citation_impact.index, year_citation_impact.values, f"Total Citations by Publication-Year Cohort - {condition}", "Total Citations", os.path.join(results_dir, f"{safe_base_name}_Total_Citations_by_Publication_Year_Cohort.jpeg"))
    if not year_annualized_impact.empty:
        pd.DataFrame({
            "Publication Year": year_annualized_impact.index.astype(int),
            "Total Annualized Citation Rate": year_annualized_impact.values,
            "Mean Annualized Citation Rate per Paper": year_mean_annualized_impact.reindex(year_annualized_impact.index).values
        }).to_csv(os.path.join(tables_dir, f"{safe_base_name}_annualized_citation_rates_by_publication_year.csv"), index=False, encoding="utf-8-sig")
        save_year_bar(year_annualized_impact.index, year_annualized_impact.values, f"Total Annualized Citation Rate by Publication Year - {condition}", "Total Annualized Citation Rate", os.path.join(results_dir, f"{safe_base_name}_Total_Annualized_Citation_Rate_by_Publication_Year.jpeg"), value_format="float")
        save_year_bar(year_mean_annualized_impact.index, year_mean_annualized_impact.values, f"Mean Annualized Citation Rate per Paper by Publication Year - {condition}", "Mean Annualized Citation Rate per Paper", os.path.join(results_dir, f"{safe_base_name}_Mean_Annualized_Citation_Rate_per_Paper_by_Publication_Year.jpeg"), value_format="float")
    open_access_indexed = open_access_status_summary.set_index("Open Access Status") if not open_access_status_summary.empty else pd.DataFrame()
    summary = {
        "Dataset": condition,
        "Records": len(df),
        "Total Citations": int(df["Cited by"].sum()),
        "Citation Annualization Reference Year": citation_reference_year if citation_reference_year is not None else "",
        "Annualization Formula": "Cited by / (reference year - publication year + 1)",
        "Entity Attribution Method": "Full counting: each paper and its citations are credited in full to every listed author and country; journals and publishers receive one attribution per paper.",
        "Open Access Comparison Type": "Unadjusted descriptive comparison; no causal effect is estimated.",
        "Publications Included in Annualized Citation Analysis": int(df["Citations per Year"].notna().sum()),
        "Total Annualized Citation Rate": float(df["Citations per Year"].sum(skipna=True)),
        "Open Access Publications": int(open_access_indexed.loc["Open Access", "Publications"]) if "Open Access" in open_access_indexed.index else 0,
        "Not Open Access Publications": int(open_access_indexed.loc["Not Open Access", "Publications"]) if "Not Open Access" in open_access_indexed.index else 0,
        "Open Access Percentage": float(open_access_indexed.loc["Open Access", "Publication Share (%)"]) if "Open Access" in open_access_indexed.index else 0,
        "Open Access Mean Cited by per Paper": float(open_access_indexed.loc["Open Access", "Mean Cited by per Paper"]) if "Open Access" in open_access_indexed.index else 0,
        "Not Open Access Mean Cited by per Paper": float(open_access_indexed.loc["Not Open Access", "Mean Cited by per Paper"]) if "Not Open Access" in open_access_indexed.index else 0,
        "Open Access Mean Annualized Citation Rate per Paper": float(open_access_indexed.loc["Open Access", "Mean Annualized Citation Rate per Paper"]) if "Open Access" in open_access_indexed.index else 0,
        "Not Open Access Mean Annualized Citation Rate per Paper": float(open_access_indexed.loc["Not Open Access", "Mean Annualized Citation Rate per Paper"]) if "Not Open Access" in open_access_indexed.index else 0,
        "Minimum Year": int(df["Year"].dropna().min()) if not df["Year"].dropna().empty else "",
        "Maximum Year": int(df["Year"].dropna().max()) if not df["Year"].dropna().empty else "",
        "Unique Authors": len(author_table),
        "Unique Countries": len(country_table),
        "Unique Journals": len(journal_table),
        "Unique Publishers": len(publisher_table),
        "Records with Extracted Authors": int(sum(bool(values) for values in author_lists)),
        "Records with Extracted Countries": int(sum(bool(values) for values in country_lists)),
        "Records with Journal": int(df["Source title"].astype(str).str.strip().ne("").sum()),
        "Records with Publisher": int(df["Publisher"].astype(str).str.strip().ne("").sum())
    }
    gc.collect()
    return summary, descriptive_statistics

GREEK_SYMBOL_NAMES = {
    "α": "alpha",
    "β": "beta",
    "γ": "gamma",
    "δ": "delta",
    "ε": "epsilon",
    "ζ": "zeta",
    "η": "eta",
    "θ": "theta",
    "ι": "iota",
    "κ": "kappa",
    "λ": "lambda",
    "μ": "mu",
    "ν": "nu",
    "ξ": "xi",
    "ο": "omicron",
    "π": "pi",
    "ρ": "rho",
    "σ": "sigma",
    "ς": "sigma",
    "τ": "tau",
    "υ": "upsilon",
    "φ": "phi",
    "χ": "chi",
    "ψ": "psi",
    "ω": "omega"
}

DISEASE_TEXT_FIELDS = ["Title", "Abstract", "Author Keywords", "Index Keywords"]
DISEASE_TOKEN_EQUIVALENTS = {
    "mutations": "mutation",
    "variations": "variation",
    "defects": "defect"
}


def normalize_disease_text(value):
    if pd.isna(value):
        return ""
    text = unicodedata.normalize("NFKC", str(value)).casefold()
    for symbol, name in GREEK_SYMBOL_NAMES.items():
        text = text.replace(symbol, f" {name} ")
    text = unicodedata.normalize("NFKD", text)
    text = "".join(character for character in text if not unicodedata.combining(character))
    text = re.sub(r"(?<=[a-z])(?=[0-9])|(?<=[0-9])(?=[a-z])", " ", text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    tokens = [DISEASE_TOKEN_EQUIVALENTS.get(token, token) for token in text.split()]
    return " ".join(tokens)


def disease_term_pattern(normalized_term):
    tokens = normalized_term.split()
    body = r"\s+".join(re.escape(token) for token in tokens)
    return re.compile(rf"(?<![a-z0-9]){body}(?![a-z0-9])")


def symbol_term_metadata(raw_term):
    text = unicodedata.normalize("NFKC", clean_text(raw_term))
    for symbol, name in GREEK_SYMBOL_NAMES.items():
        text = text.replace(symbol, name)
    decomposed = unicodedata.normalize("NFKD", text)
    ascii_text = "".join(character for character in decomposed if not unicodedata.combining(character))
    chunks = re.findall(r"[A-Za-z0-9]+", ascii_text)
    compact = "".join(chunks)
    letters = "".join(character for character in ascii_text if character.isalpha())
    all_upper = bool(letters) and letters.upper() == letters and letters.lower() != letters
    contains_digit = any(character.isdigit() for character in compact)
    context_tokens = {
        "mutation",
        "mutations",
        "variation",
        "variations",
        "defect",
        "defects",
        "deficiency",
        "deficiencies",
        "deficient",
        "syndrome",
        "disease",
        "disorder",
        "immunodeficiency",
        "haploinsufficiency",
        "gof",
        "lof",
        "gain",
        "loss",
        "function",
        "associated",
        "related",
        "type",
        "class",
        "receptor",
        "gene",
        "protein",
        "susceptibility",
        "anomaly",
        "anomalies"
    }
    has_context_token = any(chunk.casefold() in context_tokens for chunk in chunks)
    uppercase_expression = bool(chunks) and len(chunks) <= 3 and len(compact) <= 12 and all_upper and not has_context_token
    compact_numeric_symbol = len(chunks) == 1 and contains_digit and len(compact) <= 15 and not has_context_token
    compact_short_candidate = len(chunks) == 1 and 1 < len(compact) <= 6 and not has_context_token
    standalone_symbol = uppercase_expression or compact_numeric_symbol or compact_short_candidate
    short_symbol = standalone_symbol and len(compact) <= 4
    reasons = []
    if uppercase_expression:
        reasons.append("uppercase_or_acronym_expression")
    if compact_numeric_symbol:
        reasons.append("compact_alphanumeric_symbol")
    if compact_short_candidate:
        reasons.append("short_compact_term")
    return {
        "standalone_symbol": standalone_symbol,
        "short_symbol": short_symbol,
        "compact_term": compact.casefold(),
        "symbol_reason": " | ".join(reasons)
    }


def build_term_index(disease_terms):
    normalized_associations = defaultdict(lambda: defaultdict(list))
    vocabulary_rows = []
    seen_vocabulary = set()
    for disease, terms in disease_terms.items():
        values = [disease] + list(terms if isinstance(terms, list) else [])
        for raw_term in values:
            raw_term = clean_text(raw_term)
            normalized_term = normalize_disease_text(raw_term)
            if not raw_term or not normalized_term:
                continue
            raw_key = raw_term.casefold()
            if raw_key not in {value.casefold() for value in normalized_associations[normalized_term][disease]}:
                normalized_associations[normalized_term][disease].append(raw_term)
            metadata = symbol_term_metadata(raw_term)
            vocabulary_key = (disease, normalized_term, raw_key)
            if vocabulary_key in seen_vocabulary:
                continue
            seen_vocabulary.add(vocabulary_key)
            vocabulary_rows.append({
                "Disease": disease,
                "Term": raw_term,
                "Normalized Term": normalized_term,
                "Token Count": len(normalized_term.split()),
                "Standalone Symbol or Acronym": metadata["standalone_symbol"],
                "Short Standalone Symbol or Acronym": metadata["short_symbol"],
                "Compact Term": metadata["compact_term"],
                "Symbol Classification Reason": metadata["symbol_reason"]
            })
    token_frequency = Counter()
    for normalized_term in normalized_associations:
        token_frequency.update(set(normalized_term.split()))
    entries = []
    anchor_index = defaultdict(list)
    for entry_id, normalized_term in enumerate(sorted(normalized_associations)):
        tokens = normalized_term.split()
        anchor = min(set(tokens), key=lambda token: (token_frequency[token], -len(token), token))
        associations = []
        symbol_terms = []
        short_symbol_terms = []
        for disease, raw_terms in sorted(normalized_associations[normalized_term].items()):
            term_metadata = [symbol_term_metadata(raw_term) for raw_term in raw_terms]
            disease_symbol_terms = [raw_term for raw_term, metadata in zip(raw_terms, term_metadata) if metadata["standalone_symbol"]]
            disease_short_terms = [raw_term for raw_term, metadata in zip(raw_terms, term_metadata) if metadata["short_symbol"]]
            symbol_terms.extend(disease_symbol_terms)
            short_symbol_terms.extend(disease_short_terms)
            associations.append({
                "disease": disease,
                "representative_term": raw_terms[0],
                "equivalent_terms": sorted(set(raw_terms), key=lambda value: (value.casefold(), value)),
                "standalone_symbol_terms": sorted(set(disease_symbol_terms), key=lambda value: (value.casefold(), value)),
                "short_symbol_terms": sorted(set(disease_short_terms), key=lambda value: (value.casefold(), value))
            })
        mapped_diseases = [association["disease"] for association in associations]
        entry = {
            "id": entry_id,
            "normalized_term": normalized_term,
            "tokens": tokens,
            "pattern": disease_term_pattern(normalized_term),
            "term_type": "single_token_boundary" if len(tokens) == 1 else "phrase_boundary",
            "associations": associations,
            "standalone_symbol": bool(symbol_terms),
            "short_symbol": bool(short_symbol_terms),
            "standalone_symbol_terms": sorted(set(symbol_terms), key=lambda value: (value.casefold(), value)),
            "short_symbol_terms": sorted(set(short_symbol_terms), key=lambda value: (value.casefold(), value)),
            "cross_disease_collision": len(mapped_diseases) > 1,
            "mapped_diseases": mapped_diseases
        }
        entries.append(entry)
        anchor_index[anchor].append(entry_id)
    vocabulary = pd.DataFrame(vocabulary_rows)
    if not vocabulary.empty:
        disease_map = {
            normalized_term: sorted(associations)
            for normalized_term, associations in (
                (normalized_term, normalized_associations[normalized_term].keys())
                for normalized_term in normalized_associations
            )
        }
        vocabulary["Cross-Disease Collision"] = vocabulary["Normalized Term"].map(lambda value: len(disease_map[value]) > 1)
        vocabulary["Mapped Diseases"] = vocabulary["Normalized Term"].map(lambda value: " | ".join(disease_map[value]))
    else:
        vocabulary["Cross-Disease Collision"] = pd.Series(dtype=bool)
        vocabulary["Mapped Diseases"] = pd.Series(dtype=str)
    return {
        "entries": entries,
        "anchor_index": anchor_index,
        "vocabulary": vocabulary
    }


def candidate_term_entries(term_index, normalized_text):
    candidate_ids = set()
    for token in set(normalized_text.split()):
        candidate_ids.update(term_index["anchor_index"].get(token, []))
    return [term_index["entries"][entry_id] for entry_id in sorted(candidate_ids)]


def evaluate_term_match(entry, normalized_text):
    matches = list(entry["pattern"].finditer(normalized_text))
    if not matches:
        return None
    first_match = matches[0]
    return {
        "rule": "normalized_token_boundary" if entry["term_type"] == "single_token_boundary" else "normalized_phrase_boundary",
        "matched_text": first_match.group(0),
        "normalized_context": normalized_text[max(0, first_match.start() - 80):min(len(normalized_text), first_match.end() + 80)],
        "match_start": first_match.start(),
        "match_end": first_match.end(),
        "match_positions": [(match.start(), match.end()) for match in matches]
    }


def aggregate_values(values):
    return " | ".join(sorted(set(clean_text(value) for value in values if clean_text(value)), key=lambda value: (value.casefold(), value)))


def disease_context_supported(normalized_text, match_start, match_end, window=8):
    context_tokens = {
        "deficiency",
        "deficient",
        "syndrome",
        "disease",
        "disorder",
        "immunodeficiency",
        "mutation",
        "variation",
        "variant",
        "defect",
        "pathogenic",
        "causative",
        "biallelic",
        "monoallelic",
        "autosomal",
        "recessive",
        "dominant",
        "linked",
        "gof",
        "lof",
        "gain",
        "loss",
        "function",
        "associated",
        "related",
        "diagnosis",
        "diagnosed",
        "patient",
        "patients",
        "case",
        "cases",
        "cohort"
    }
    tokens = normalized_text.split()
    start_token = len(normalized_text[:match_start].split())
    end_token = len(normalized_text[:match_end].split())
    nearby = tokens[max(0, start_token - window):min(len(tokens), end_token + window)]
    return bool(context_tokens.intersection(nearby))


def numeric_distribution(values):
    numeric = pd.to_numeric(pd.Series(values), errors="coerce").replace([np.inf, -np.inf], np.nan).dropna()
    if numeric.empty:
        return {
            "Count": 0,
            "Mean": 0.0,
            "Median": 0.0,
            "Standard Deviation": 0.0,
            "Q1": 0.0,
            "Q3": 0.0,
            "IQR": 0.0,
            "Minimum": 0.0,
            "Maximum": 0.0
        }
    q1 = float(numeric.quantile(0.25))
    q3 = float(numeric.quantile(0.75))
    return {
        "Count": int(numeric.count()),
        "Mean": float(numeric.mean()),
        "Median": float(numeric.median()),
        "Standard Deviation": float(numeric.std(ddof=1)) if len(numeric) > 1 else 0.0,
        "Q1": q1,
        "Q3": q3,
        "IQR": q3 - q1,
        "Minimum": float(numeric.min()),
        "Maximum": float(numeric.max())
    }


def add_distribution_columns(row, prefix, values):
    statistics = numeric_distribution(values)
    names = {
        "Count": f"{prefix} Observations",
        "Mean": f"Mean {prefix}",
        "Median": f"Median {prefix}",
        "Standard Deviation": f"Standard Deviation of {prefix}",
        "Q1": f"Q1 {prefix}",
        "Q3": f"Q3 {prefix}",
        "IQR": f"IQR {prefix}",
        "Minimum": f"Minimum {prefix}",
        "Maximum": f"Maximum {prefix}"
    }
    for metric, value in statistics.items():
        row[names[metric]] = value


def disease_descriptive_statistics(summary, assignment_table):
    rows = []

    def add_row(unit, metric, values):
        statistics = numeric_distribution(values)
        row = {"Unit": unit, "Metric": metric}
        row.update(statistics)
        rows.append(row)

    diseases_with_papers = summary[summary["Total Papers"] > 0].copy()
    disease_metrics = [
        "Total Papers",
        "Total Citations",
        "Mean Citations per Paper",
        "Median Citations per Paper",
        "Total Annualized Citation Rate",
        "Mean Annualized Citation Rate per Paper",
        "Normalized Papers per Year",
        "Active Publication Years",
        "Publication Window (Years)",
        "Strict High-Confidence Retention (%)"
    ]
    for metric in disease_metrics:
        if metric in diseases_with_papers.columns:
            add_row("Disease categories with at least one paper", metric, diseases_with_papers[metric])
    if not assignment_table.empty:
        unique_records = assignment_table.drop_duplicates("Original CSV Row")
        add_row("Unique records assigned to at least one disease", "Citations per Paper", unique_records["Cited by"])
        add_row("Unique records assigned to at least one disease", "Annualized Citation Rate per Paper", unique_records["Citations per Year"])
        add_row("Disease-paper assignments", "Citations per Paper", assignment_table["Cited by"])
        add_row("Disease-paper assignments", "Annualized Citation Rate per Paper", assignment_table["Citations per Year"])
    return pd.DataFrame(rows)


def format_disease_workbook(writer, frames):
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet_name, frame in frames.items():
        worksheet = writer.book[sheet_name]
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.row_dimensions[1].height = 30
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_index, column_name in enumerate(frame.columns, start=1):
            values = [str(column_name)] + [clean_text(value) for value in frame[column_name].head(2000)]
            width = max(len(value) for value in values) + 2 if values else 12
            if sheet_name == "Methodology_and_Limits" and column_name == "Value":
                width = 100
            elif column_name in {"Disease", "Metric", "Item", "Unit"}:
                width = min(max(width, 22), 45)
            else:
                width = min(max(width, 12), 24)
            worksheet.column_dimensions[get_column_letter(column_index)].width = width
            for cell in worksheet[get_column_letter(column_index)][1:]:
                cell.alignment = Alignment(vertical="top", wrap_text=sheet_name == "Methodology_and_Limits")
                if isinstance(cell.value, float):
                    cell.number_format = "0.00"
                if "%" in str(column_name):
                    cell.number_format = "0.00"
        if sheet_name == "Methodology_and_Limits":
            for row in worksheet.iter_rows(min_row=2):
                worksheet.row_dimensions[row[0].row].height = 45


def analyze_disease_occurrences(combined_df, disease_terms, output_dir, citation_reference_year=None):
    disease_dir = os.path.join(output_dir, "specific_iei_diseases")
    os.makedirs(disease_dir, exist_ok=True)
    for filename in os.listdir(disease_dir):
        path = os.path.join(disease_dir, filename)
        if os.path.isfile(path) or os.path.islink(path):
            os.remove(path)
    legacy_summary_path = os.path.join(output_dir, "iei_disease_analysis_summary.csv")
    if os.path.exists(legacy_summary_path):
        os.remove(legacy_summary_path)
    combined = combined_df.copy()
    combined["Year"] = pd.to_numeric(combined["Year"], errors="coerce")
    combined, citation_reference_year = add_annualized_citation_metrics(combined, citation_reference_year)
    combined["Original CSV Row"] = pd.to_numeric(combined["Original CSV Row"], errors="coerce")
    if combined["Original CSV Row"].isna().any():
        raise ValueError("Original CSV Row contains invalid values.")
    combined["Original CSV Row"] = combined["Original CSV Row"].astype(int)
    term_index = build_term_index(disease_terms)
    vocabulary = term_index["vocabulary"]
    high_specificity_fields = {"Title", "Author Keywords", "Index Keywords"}
    accepted_assignments = defaultdict(dict)
    rejection_counts = Counter()
    candidate_pair_count = 0
    candidate_evidence_count = 0
    record_metadata = {}
    for record_id, (_, row) in enumerate(combined.iterrows(), start=1):
        original_csv_row = int(row["Original CSV Row"])
        year = int(row["Year"]) if pd.notna(row["Year"]) else np.nan
        metadata = {
            "Original CSV Row": original_csv_row,
            "Record ID": record_id,
            "Title": clean_text(row.get("Title", "")),
            "DOI": clean_text(row.get("DOI", "")),
            "Year": year,
            "Cited by": int(row["Cited by"]),
            "Citation Years": float(row["Citation Years"]) if pd.notna(row["Citation Years"]) else np.nan,
            "Citations per Year": float(row["Citations per Year"]) if pd.notna(row["Citations per Year"]) else np.nan
        }
        record_metadata[original_csv_row] = metadata
        evidence_by_disease = defaultdict(list)
        for field_name in DISEASE_TEXT_FIELDS:
            normalized_text = normalize_disease_text(row.get(field_name, ""))
            if not normalized_text:
                continue
            for entry in candidate_term_entries(term_index, normalized_text):
                match = evaluate_term_match(entry, normalized_text)
                if not match:
                    continue
                for association in entry["associations"]:
                    non_symbol_terms = [
                        term
                        for term in association["equivalent_terms"]
                        if not symbol_term_metadata(term)["standalone_symbol"]
                    ]
                    evidence_by_disease[association["disease"]].append({
                        "Matched Term": association["representative_term"],
                        "Equivalent Terms": association["equivalent_terms"],
                        "Normalized Term": entry["normalized_term"],
                        "Token Count": len(entry["tokens"]),
                        "Compact Length": len(entry["normalized_term"].replace(" ", "")),
                        "Field": field_name,
                        "Symbol Only": not bool(non_symbol_terms),
                        "Cross-Disease Collision": entry["cross_disease_collision"],
                        "Context Supported": any(
                            disease_context_supported(normalized_text, match_start, match_end)
                            for match_start, match_end in match["match_positions"]
                        )
                    })
                    candidate_evidence_count += 1
        candidate_pair_count += len(evidence_by_disease)
        for disease, evidence in evidence_by_disease.items():
            accepted_evidence = list(evidence)
            phrase_evidence = [item for item in accepted_evidence if not item["Symbol Only"]]
            symbol_evidence = [item for item in accepted_evidence if item["Symbol Only"]]
            all_fields = sorted(set(item["Field"] for item in accepted_evidence))
            accepted = False
            strict_high_confidence = False
            evidence_category = ""
            if phrase_evidence:
                accepted = True
                evidence_category = "Normalized disease phrase"
                phrase_fields = set(item["Field"] for item in phrase_evidence)
                phrase_terms = defaultdict(set)
                for item in phrase_evidence:
                    phrase_terms[item["Normalized Term"]].add(item["Field"])
                strict_high_confidence = any(
                    item["Field"] in high_specificity_fields or item["Token Count"] >= 3
                    for item in phrase_evidence
                ) or any(len(fields) >= 2 for fields in phrase_terms.values())
            elif symbol_evidence:
                symbol_fields = set(item["Field"] for item in symbol_evidence)
                contextual_support = any(item["Context Supported"] for item in symbol_evidence)
                high_specificity_support = any(item["Field"] in high_specificity_fields for item in symbol_evidence)
                repeated_field_support = len(symbol_fields) >= 2
                sufficient_length = max(item["Compact Length"] for item in symbol_evidence) >= 4
                if contextual_support and (high_specificity_support or repeated_field_support):
                    accepted = True
                    evidence_category = "Context-supported standalone symbol or acronym"
                elif repeated_field_support and sufficient_length:
                    accepted = True
                    evidence_category = "Repeated-field standalone symbol or acronym"
                strict_high_confidence = accepted and contextual_support and high_specificity_support and repeated_field_support
            if accepted:
                accepted_assignments[original_csv_row][disease] = {
                    "Strict High-Confidence": bool(strict_high_confidence),
                    "Evidence Category": evidence_category,
                    "Matched Terms": aggregate_values(item["Matched Term"] for item in accepted_evidence),
                    "Match Fields": aggregate_values(all_fields)
                }
            elif accepted_evidence:
                rejection_counts["Unsupported standalone symbol or acronym"] += 1
    assignment_rows = []
    for original_csv_row, diseases in sorted(accepted_assignments.items()):
        metadata = record_metadata[original_csv_row]
        for disease, evidence in sorted(diseases.items()):
            assignment_rows.append({
                **metadata,
                "Disease": disease,
                "Strict High-Confidence": evidence["Strict High-Confidence"],
                "Evidence Category": evidence["Evidence Category"],
                "Matched Terms": evidence["Matched Terms"],
                "Match Fields": evidence["Match Fields"]
            })
    assignment_columns = [
        "Original CSV Row",
        "Record ID",
        "Title",
        "DOI",
        "Year",
        "Cited by",
        "Citation Years",
        "Citations per Year",
        "Disease",
        "Strict High-Confidence",
        "Evidence Category",
        "Matched Terms",
        "Match Fields"
    ]
    assignment_table = pd.DataFrame(assignment_rows, columns=assignment_columns)
    summary_rows = []
    disease_results_for_plot = {}
    annual_rows = []
    for disease in disease_terms:
        subset = assignment_table[assignment_table["Disease"] == disease].copy() if not assignment_table.empty else pd.DataFrame(columns=assignment_columns)
        citations = pd.to_numeric(subset["Cited by"], errors="coerce").fillna(0)
        normalized_citations = pd.to_numeric(subset["Citations per Year"], errors="coerce").dropna()
        years = pd.to_numeric(subset["Year"], errors="coerce").dropna().astype(int)
        papers_with_year = int(years.count())
        first_year = int(years.min()) if papers_with_year else np.nan
        last_year = int(years.max()) if papers_with_year else np.nan
        publication_window = last_year - first_year + 1 if papers_with_year else 0
        annual_counts = []
        papers_per_year = defaultdict(int)
        citation_impact_by_year = defaultdict(int)
        normalized_impact_by_year = defaultdict(float)
        if papers_with_year:
            for year in range(first_year, last_year + 1):
                year_subset = subset[pd.to_numeric(subset["Year"], errors="coerce") == year]
                year_citations = pd.to_numeric(year_subset["Cited by"], errors="coerce").fillna(0)
                year_normalized = pd.to_numeric(year_subset["Citations per Year"], errors="coerce").dropna()
                papers = int(len(year_subset))
                annual_counts.append(papers)
                papers_per_year[year] = papers
                citation_impact_by_year[year] = int(year_citations.sum())
                normalized_impact_by_year[year] = float(year_normalized.sum())
                annual_rows.append({
                    "Disease": disease,
                    "Year": year,
                    "Papers": papers,
                    "Share of Disease Papers with Year (%)": papers / papers_with_year * 100 if papers_with_year else 0,
                    "Total Citations": int(year_citations.sum()),
                    "Mean Citations per Paper": float(year_citations.mean()) if papers else 0,
                    "Median Citations per Paper": float(year_citations.median()) if papers else 0,
                    "Total Annualized Citation Rate": float(year_normalized.sum()),
                    "Mean Annualized Citation Rate per Paper": float(year_normalized.mean()) if not year_normalized.empty else 0
                })
        strict_subset = subset[subset["Strict High-Confidence"]].copy() if not subset.empty else subset.copy()
        strict_papers = int(len(strict_subset))
        strict_citations = pd.to_numeric(strict_subset["Cited by"], errors="coerce").fillna(0)
        strict_normalized = pd.to_numeric(strict_subset["Citations per Year"], errors="coerce").dropna()
        total_papers = int(len(subset))
        row = {
            "Disease": disease,
            "Total Papers": total_papers,
            "Papers with Year": papers_with_year,
            "Papers without Year": total_papers - papers_with_year,
            "Total Citations": int(citations.sum()) if total_papers else 0,
            "Papers with at Least One Citation": int((citations > 0).sum()) if total_papers else 0,
            "Cited Papers (%)": float((citations > 0).mean() * 100) if total_papers else 0,
            "Papers Included in Annualized Citation Analysis": int(normalized_citations.count()),
            "Total Annualized Citation Rate": float(normalized_citations.sum()),
            "First Publication Year": first_year,
            "Last Publication Year": last_year,
            "Publication Window (Years)": publication_window,
            "Active Publication Years": int(years.nunique()) if papers_with_year else 0,
            "Papers per Active Publication Year": papers_with_year / years.nunique() if papers_with_year else 0,
            "Normalized Papers per Year": papers_with_year / publication_window if publication_window else 0,
            "Strict High-Confidence Papers": strict_papers,
            "Strict High-Confidence Retention (%)": strict_papers / total_papers * 100 if total_papers else 0,
            "Strict High-Confidence Total Citations": int(strict_citations.sum()) if strict_papers else 0,
            "Strict High-Confidence Total Annualized Citation Rate": float(strict_normalized.sum()),
            "Phrase-Supported Papers": int((subset["Evidence Category"] == "Normalized disease phrase").sum()) if total_papers else 0,
            "Contextual or Repeated Symbol-Supported Papers": int((subset["Evidence Category"] != "Normalized disease phrase").sum()) if total_papers else 0
        }
        add_distribution_columns(row, "Citations per Paper", citations)
        add_distribution_columns(row, "Annualized Citation Rate per Paper", normalized_citations)
        add_distribution_columns(row, "Annual Papers", annual_counts)
        summary_rows.append(row)
        disease_results_for_plot[disease] = {
            "total_papers": total_papers,
            "papers_per_year": papers_per_year,
            "citation_impact_by_publication_year": citation_impact_by_year,
            "total_annualized_citation_rate_by_publication_year": normalized_impact_by_year
        }
    summary = pd.DataFrame(summary_rows).sort_values(
        ["Total Papers", "Total Citations", "Disease"],
        ascending=[False, False, True]
    )
    summary["First Publication Year"] = summary["First Publication Year"].astype("Int64")
    summary["Last Publication Year"] = summary["Last Publication Year"].astype("Int64")
    annual_trends = pd.DataFrame(annual_rows)
    if not annual_trends.empty:
        annual_trends = annual_trends.sort_values(["Disease", "Year"])
    descriptive_statistics = disease_descriptive_statistics(summary, assignment_table)
    collision_terms = int(vocabulary.loc[vocabulary["Cross-Disease Collision"], "Normalized Term"].nunique()) if not vocabulary.empty else 0
    raw_dictionary_terms = int(len(vocabulary))
    unique_normalized_terms = int(vocabulary["Normalized Term"].nunique()) if not vocabulary.empty else 0
    accepted_assignment_count = int(len(assignment_table))
    strict_assignment_count = int(assignment_table["Strict High-Confidence"].sum()) if not assignment_table.empty else 0
    assigned_records = int(assignment_table["Original CSV Row"].nunique()) if not assignment_table.empty else 0
    phrase_assignments = int((assignment_table["Evidence Category"] == "Normalized disease phrase").sum()) if not assignment_table.empty else 0
    symbol_assignments = accepted_assignment_count - phrase_assignments
    limitation_text = (
        "Disease assignments were generated by a conservative rule-based dictionary matcher and were not manually adjudicated. "
        "Exact normalized token boundaries and contextual or repeated-field requirements for standalone symbols and acronyms reduce false-positive matching but cannot eliminate it. Shared terms may assign a publication to multiple related disease categories when the curated dictionary maps the term to each category. "
        "Residual false-positive and false-negative classification may remain because of incomplete titles, abstracts or keywords, acronym ambiguity, evolving IEI nomenclature, imperfect source synonyms, and papers that discuss a disease without making it their principal focus. "
        "Disease-level estimates should therefore be interpreted as automated bibliometric approximations rather than clinically validated classifications."
    )
    methodology_rows = [
        {"Section": "Method", "Item": "Text fields searched", "Value": "Title; Abstract; Author Keywords; Index Keywords"},
        {"Section": "Method", "Item": "Dictionary source", "Value": f"User-supplied curated IEI disease and synonym dictionary containing {len(disease_terms)} canonical disease categories, {raw_dictionary_terms} disease-term rows, and {unique_normalized_terms} unique normalized terms."},
        {"Section": "Method", "Item": "Text normalization", "Value": "Unicode normalization, case folding, diacritic removal, Greek-symbol expansion, punctuation removal, letter-number boundary separation, whitespace normalization, and harmonization of mutation/mutations, variation/variations, and defect/defects."},
        {"Section": "Method", "Item": "Primary matching rule", "Value": "Exact normalized token-boundary matching only. Fuzzy matching and unrestricted substring matching are not used."},
        {"Section": "Interpretation", "Item": "Cross-disease term collisions", "Value": "A normalized term mapped to more than one disease is evaluated for every mapped disease and may support assignment to each category. Related parent, subtype, and alternative-name categories may therefore overlap."},
        {"Section": "False-positive control", "Item": "Standalone symbols and acronyms", "Value": "Every exact-boundary occurrence of a standalone symbol or acronym is evaluated for nearby disease-related context. The term is context-supported when any occurrence satisfies the local contextual rule and is accepted only when that support occurs in a high-specificity field or alongside evidence in multiple fields, or when a sufficiently long acronym is repeated across multiple fields."},
        {"Section": "False-positive control", "Item": "Disease phrases", "Value": "Non-symbol disease names, synonyms, and gene-plus-pathogenic-context phrases are accepted after exact normalized boundary matching."},
        {"Section": "Robustness", "Item": "Strict high-confidence sensitivity subset", "Value": "A stricter subset is identified when phrase evidence occurs in the title or keywords, contains at least three normalized tokens, is repeated across fields, or when contextual symbol evidence is present in both a high-specificity field and another field."},
        {"Section": "Interpretation", "Item": "Multiple disease assignments", "Value": "A publication may be assigned to more than one disease when accepted evidence is present for each disease, including when a curated shared term maps to multiple related categories."},
        {"Section": "Validation", "Item": "Manual validation performed", "Value": "No"},
        {"Section": "Run summary", "Item": "Combined deduplicated records scanned", "Value": len(combined)},
        {"Section": "Run summary", "Item": "Records assigned to at least one disease", "Value": assigned_records},
        {"Section": "Run summary", "Item": "Accepted record-disease assignments", "Value": accepted_assignment_count},
        {"Section": "Run summary", "Item": "Strict high-confidence assignments", "Value": strict_assignment_count},
        {"Section": "Run summary", "Item": "Strict high-confidence retention (%)", "Value": strict_assignment_count / accepted_assignment_count * 100 if accepted_assignment_count else 0},
        {"Section": "Run summary", "Item": "Phrase-supported assignments", "Value": phrase_assignments},
        {"Section": "Run summary", "Item": "Contextual or repeated symbol-supported assignments", "Value": symbol_assignments},
        {"Section": "Run summary", "Item": "Candidate record-disease pairs before filtering", "Value": candidate_pair_count},
        {"Section": "Run summary", "Item": "Candidate term-evidence occurrences before filtering", "Value": candidate_evidence_count},
        {"Section": "Run summary", "Item": "Rejected unsupported standalone symbol or acronym pairs", "Value": rejection_counts["Unsupported standalone symbol or acronym"]},
        {"Section": "Run summary", "Item": "Cross-disease normalized terms in dictionary", "Value": collision_terms},
        {"Section": "Run summary", "Item": "Citation annualization reference year", "Value": citation_reference_year if citation_reference_year is not None else ""},
        {"Section": "Discussion limitation", "Item": "Recommended limitation statement", "Value": limitation_text}
    ]
    methodology = pd.DataFrame(methodology_rows)
    plot_top_disease_trends(
        disease_results_for_plot,
        os.path.join(disease_dir, "top_20_diseases_trends.jpeg"),
        top_n=20
    )
    workbook_path = os.path.join(disease_dir, "disease_analysis.xlsx")
    frames = {
        "Disease_Summary": summary,
        "Annual_Trends": annual_trends,
        "Descriptive_Statistics": descriptive_statistics,
        "Methodology_and_Limits": methodology
    }
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        for sheet_name, frame in frames.items():
            frame.to_excel(writer, sheet_name=sheet_name, index=False)
        format_disease_workbook(writer, frames)
    return {
        "Total diseases analyzed": len(disease_terms),
        "Diseases with papers": int((summary["Total Papers"] > 0).sum()),
        "Combined deduplicated records": len(combined),
        "Citation annualization reference year": citation_reference_year if citation_reference_year is not None else "",
        "Accepted record-disease assignments": accepted_assignment_count,
        "Strict high-confidence assignments": strict_assignment_count,
        "Records assigned to at least one disease": assigned_records,
        "Manual validation performed": False
    }

def generate_extended_colormap(num_colors):
    colormaps = ["tab20", "tab20b", "tab20c", "Set3", "Set1", "Set2", "Dark2", "Paired"]
    colors = []
    for cmap_name in colormaps:
        cmap = plt.cm.get_cmap(cmap_name)
        colors.extend([cmap(i) for i in range(cmap.N)])
    seen = set()
    unique_colors = []
    for color in colors:
        key = tuple(color)
        if key not in seen:
            seen.add(key)
            unique_colors.append(color)
    if len(unique_colors) < num_colors:
        additional = num_colors - len(unique_colors)
        for i in range(additional):
            unique_colors.append(plt.cm.hsv(i / max(additional, 1)))
    return unique_colors[:num_colors]


def plot_all_disease_trends(disease_results, output_dir):
    diseases_with_data = {disease: data for disease, data in disease_results.items() if data["total_papers"] > 0}
    if not diseases_with_data:
        return
    sorted_diseases = dict(sorted(diseases_with_data.items(), key=lambda x: x[1]["total_papers"], reverse=True))
    colors = generate_extended_colormap(len(sorted_diseases))
    plt.figure(figsize=(16, 10))
    color_data = []
    for i, (disease, data) in enumerate(sorted_diseases.items()):
        years = sorted(data["papers_per_year"].keys())
        papers = [data["papers_per_year"][year] for year in years]
        plt.plot(years, papers, marker="o", linewidth=1.5, markersize=4, label=disease, color=colors[i], alpha=0.8)
        rgb = tuple(int(x * 255) for x in colors[i][:3])
        color_data.append({"Disease": disease, "Red": rgb[0], "Green": rgb[1], "Blue": rgb[2], "Total_Papers": data["total_papers"]})
    plt.xlabel("Year", fontsize=14)
    plt.ylabel("Number of Research Papers", fontsize=14)
    plt.title(f"Research Papers per Year by Disease (All {len(sorted_diseases)} diseases)", fontsize=16, fontweight="bold")
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, "all_diseases_trends.jpeg"), dpi=600, bbox_inches="tight", format="jpeg")
    plt.close()


def plot_top_disease_trends(disease_results, filename, top_n=20):
    diseases_with_data = {disease: data for disease, data in disease_results.items() if data["total_papers"] > 0}
    if not diseases_with_data:
        return
    sorted_diseases = sorted(diseases_with_data.items(), key=lambda x: x[1]["total_papers"], reverse=True)[:top_n]
    colors = plt.cm.tab20(np.linspace(0, 1, len(sorted_diseases)))
    plt.figure(figsize=(14, 8))
    color_data = []
    for i, (disease, data) in enumerate(sorted_diseases):
        observed_years = sorted(data["papers_per_year"].keys())
        years = list(range(min(observed_years), max(observed_years) + 1)) if observed_years else []
        papers = [data["papers_per_year"].get(year, 0) for year in years]
        plt.plot(years, papers, marker="o", linewidth=2.5, markersize=6, label=f"{disease} ({data['total_papers']} papers)", color=colors[i], alpha=0.9)
        rgb = tuple(int(x * 255) for x in colors[i][:3])
        color_data.append({"Disease": disease, "Red": rgb[0], "Green": rgb[1], "Blue": rgb[2], "Total_Papers": data["total_papers"]})
    plt.xlabel("Year", fontsize=14)
    plt.ylabel("Number of Research Papers", fontsize=14)
    plt.title(f"Research Papers per Year by Disease (Top {len(sorted_diseases)} Diseases)", fontsize=16, fontweight="bold")
    plt.grid(True, alpha=0.3)
    plt.legend(bbox_to_anchor=(1.05, 1), loc="upper left", fontsize=10)
    plt.tight_layout()
    plt.savefig(filename, dpi=300, bbox_inches="tight", format="jpeg")
    plt.close()


def save_combined_category_trends(pid_df, iei_df, output_dir, citation_reference_year=None):
    results_dir = os.path.join(output_dir, "bibliometric_results")
    tables_dir = os.path.join(output_dir, "bibliometric_tables")
    os.makedirs(results_dir, exist_ok=True)
    os.makedirs(tables_dir, exist_ok=True)
    remove_existing_files([
        os.path.join(results_dir, "PID_IEI_Annual_Publication_Comparison.jpeg"),
        os.path.join(results_dir, "PID_IEI_Average_Citations_per_Year_per_Paper_by_Publication_Year_Comparison.jpeg"),
        os.path.join(results_dir, "PID_IEI_Cumulative_Citation_Impact_by_Publication_Year_Comparison.jpeg"),
        os.path.join(results_dir, "PID_IEI_Normalized_Citation_Impact_by_Publication_Year_Comparison.jpeg"),
        os.path.join(tables_dir, "pid_iei_annual_publications_cumulative_citation_impact_comparison.csv"),
        os.path.join(tables_dir, "pid_iei_annual_publications_cumulative_and_normalized_citation_impact_comparison.csv")
    ])
    pid_df, _ = add_annualized_citation_metrics(pid_df, citation_reference_year)
    iei_df, _ = add_annualized_citation_metrics(iei_df, citation_reference_year)
    pid_publications, pid_citation_impact = annual_publication_and_citation_impact_tables(pid_df)
    iei_publications, iei_citation_impact = annual_publication_and_citation_impact_tables(iei_df)
    pid_annualized_impact, pid_mean_annualized_impact = annualized_citation_rate_tables(pid_df)
    iei_annualized_impact, iei_mean_annualized_impact = annualized_citation_rate_tables(iei_df)
    years = sorted(set(pid_publications.index.astype(int).tolist()) | set(iei_publications.index.astype(int).tolist()) | set(pid_citation_impact.index.astype(int).tolist()) | set(iei_citation_impact.index.astype(int).tolist()) | set(pid_annualized_impact.index.astype(int).tolist()) | set(iei_annualized_impact.index.astype(int).tolist()))
    rows = []
    for year in years:
        rows.append({
            "Year": year,
            "Primary Immunodeficiency Publications": int(pid_publications.get(year, 0)),
            "Inborn Errors of Immunity Publications": int(iei_publications.get(year, 0)),
            "Primary Immunodeficiency Total Citations by Publication-Year Cohort": int(pid_citation_impact.get(year, 0)),
            "Inborn Errors of Immunity Total Citations by Publication-Year Cohort": int(iei_citation_impact.get(year, 0)),
            "Primary Immunodeficiency Total Annualized Citation Rate": float(pid_annualized_impact.get(year, 0)),
            "Inborn Errors of Immunity Total Annualized Citation Rate": float(iei_annualized_impact.get(year, 0)),
            "Primary Immunodeficiency Mean Annualized Citation Rate per Paper": float(pid_mean_annualized_impact.get(year, 0)),
            "Inborn Errors of Immunity Mean Annualized Citation Rate per Paper": float(iei_mean_annualized_impact.get(year, 0))
        })
    pd.DataFrame(rows).to_csv(os.path.join(tables_dir, "pid_iei_annual_bibliometric_metrics_comparison.csv"), index=False, encoding="utf-8-sig")

def resolve_combined_preprocessed_path(pid_path, iei_path, combined_path):
    if combined_path:
        return combined_path
    pid_dir = os.path.dirname(os.path.abspath(pid_path))
    iei_dir = os.path.dirname(os.path.abspath(iei_path))
    if pid_dir != iei_dir:
        raise ValueError(
            'The PID and IEI preprocessed files are in different directories. '
            'Provide --combined-preprocessed with the combined preprocessing output.'
        )
    return os.path.join(pid_dir, 'combined_pid_iei_preprocessed_deduplicated_by_title_year.csv')


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--pid", default=os.path.join("preprocessed_bibliometric_data", "primary_immunodeficiency_diseases_preprocessed.csv"))
    parser.add_argument("--iei", default=os.path.join("preprocessed_bibliometric_data", "inborn_errors_of_immunity_preprocessed.csv"))
    parser.add_argument("--disease-terms", default="iei_disease_terms.json")
    parser.add_argument("--combined-preprocessed", default="")
    parser.add_argument("--out", default="bibliometric_analysis_results")
    parser.add_argument("--stage", default="all", choices=["all", "dataset", "combined", "disease"])
    parser.add_argument("--dataset-name", default="")
    parser.add_argument("--citation-reference-year", type=int, default=None)
    args = parser.parse_args()
    os.makedirs(args.out, exist_ok=True)
    pid_df = read_preprocessed(args.pid)
    iei_df = read_preprocessed(args.iei)
    valid_years = pd.concat([pid_df["Year"], iei_df["Year"]], ignore_index=True).dropna()
    citation_reference_year = args.citation_reference_year if args.citation_reference_year is not None else (int(valid_years.max()) if not valid_years.empty else None)
    if citation_reference_year is not None and not valid_years.empty and citation_reference_year < int(valid_years.max()):
        raise ValueError("The citation reference year cannot be earlier than the latest publication year in the input datasets.")
    summaries = []
    descriptive_statistics_tables = []
    combined_df = None
    if args.stage in ["all", "dataset"]:
        if args.dataset_name in ["", "primary_immunodeficiency_diseases"]:
            summary, descriptive_statistics = analyze_dataset(pid_df, "primary_immunodeficiency_diseases", args.out, citation_reference_year)
            summaries.append(summary)
            descriptive_statistics_tables.append(descriptive_statistics)
        if args.dataset_name in ["", "inborn_errors_of_immunity"]:
            summary, descriptive_statistics = analyze_dataset(iei_df, "inborn_errors_of_immunity", args.out, citation_reference_year)
            summaries.append(summary)
            descriptive_statistics_tables.append(descriptive_statistics)
    if args.stage in ["all", "combined"]:
        save_combined_category_trends(pid_df, iei_df, args.out, citation_reference_year)
        combined_path = resolve_combined_preprocessed_path(args.pid, args.iei, args.combined_preprocessed)
        if not os.path.exists(combined_path):
            raise FileNotFoundError(
                f'Combined preprocessed file not found: {combined_path}. '
                'Run preprocess_scopus_bibliometrics.py before the combined network analysis.'
            )
        combined_df = read_preprocessed(combined_path)
        save_combined_network_analysis(combined_df, args.out, citation_reference_year)
    if args.stage in ["all", "disease"]:
        if combined_df is None:
            combined_path = resolve_combined_preprocessed_path(args.pid, args.iei, args.combined_preprocessed)
            if not os.path.exists(combined_path):
                raise FileNotFoundError(
                    f'Combined preprocessed file not found: {combined_path}. '
                    'Run preprocess_scopus_bibliometrics.py before the disease analysis.'
                )
            combined_df = read_preprocessed(combined_path)
        with open(args.disease_terms, "r", encoding="utf-8") as handle:
            disease_terms = json.load(handle)
        analyze_disease_occurrences(combined_df, disease_terms, args.out, citation_reference_year)
    if summaries:
        pd.DataFrame(summaries).to_csv(os.path.join(args.out, "bibliometric_dataset_summary.csv"), index=False, encoding="utf-8-sig")
    if descriptive_statistics_tables:
        tables_dir = os.path.join(args.out, "bibliometric_tables")
        os.makedirs(tables_dir, exist_ok=True)
        pd.concat(descriptive_statistics_tables, ignore_index=True).to_csv(os.path.join(tables_dir, "bibliometric_descriptive_statistics.csv"), index=False, encoding="utf-8-sig")
    print("Analysis complete")
    print(os.path.abspath(args.out))


if __name__ == "__main__":
    main()
